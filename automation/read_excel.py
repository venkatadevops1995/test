import openpyxl
import re
import datetime
import sys
import os

wts_file = 'vedika_dl_output_samples_unique_emp_names_DVR_Kiran_Durga_automation_100.xlsx'
wts_file = os.path.join(os.path.dirname(__file__),wts_file)
# wts_file = '/home/rakumar/TEST_ATDATA/automation/vedika_dl_output_samples_unique_emp_names_DVR_Kiran_Durga_automation.xlsx'
# WTS_DATA = []

def get_email_id():
    EMP_EMAIL = {}
    MANAGER_DATA = {}
    wts_wb = openpyxl.load_workbook(wts_file)
    wts_ws = wts_wb[wts_wb.sheetnames[0]]
    wts_max_row = wts_ws.max_row
    wts_max_col = wts_ws.max_column
    name_col = 0
    email_col = 0
    manager_col = 0
    # wts_max_row = 2
    for i in range(1, wts_max_col + 1):
        wts_cell = wts_ws.cell(row = 1, column = i)
        if(wts_cell.value == 'Name'): 
            name_col = i
            continue
        elif(wts_cell.value == 'Email'):
            email_col = i
        elif(wts_cell.value == 'Reporting Manager'):
            manager_col = i
    for j in range(2, wts_max_row + 1):
        name = wts_ws.cell(row = j, column = name_col).value
        email = wts_ws.cell(row = j, column = email_col).value
        manager = wts_ws.cell(row = j, column = manager_col).value
        if(name != None and email != None):
            EMP_EMAIL.update({name:email})
        if(name != None and manager != None ):
            if(manager in MANAGER_DATA):
                MANAGER_DATA[manager].append(name)
                # print(MANAGER_DATA)
            else:
                MANAGER_DATA.update({manager:[name]})
    return EMP_EMAIL,MANAGER_DATA

def get_wts_data():
    WTS_DATA = []
    SUBMIT_INFO = {}
    ACTION_INFO = {}
    wts_wb = openpyxl.load_workbook(wts_file)
    # print(wts_wb.sheetnames)
    #   sys.exit()
    wts_ws = wts_wb[wts_wb.sheetnames[1]]
    wts_max_row = wts_ws.max_row
    wts_max_col = 9
    emp_data = []
    for j in range(3, wts_max_row + 1):
        proj_wts = []
        for i in range(1, wts_max_col + 1):
            wts_cell = wts_ws.cell(row = j, column = i)
            # print(i,wts_cell.value)
            if(i==1):
                if(wts_cell.value!=None):
                    if(len(emp_data)!=0):
                        WTS_DATA.append(emp_data)
                    emp_data = []
                    emp_data.append(wts_cell.value.strip())
                    #print(wts_ws.cell(row = j, column = 10).value,wts_ws.cell(row = j, column = 11).value)
                    SUBMIT_INFO.update({ wts_ws.cell(row = j, column = 1).value:wts_ws.cell(row = j, column = 10).value})
                    ACTION_INFO.update({ wts_ws.cell(row = j, column = 1).value:wts_ws.cell(row = j, column = 11).value})
                continue
            elif(i==2):
                if(wts_cell.value == None):
                    continue
                emp_data.append(wts_cell.value.strip())
                proj_wts = []
                continue
            elif(i<=9):
                proj_wts.append(wts_cell.value)
            if(i==9 and wts_ws.cell(row = j, column = 2).value != None ):
                # if len([x for x in proj_wts if x!=None]):
                emp_data.append(proj_wts)
    if(len(emp_data)!=0):
        WTS_DATA.append(emp_data)
    return WTS_DATA,SUBMIT_INFO,ACTION_INFO

# def get_manager_dashboard_data():
#     WTS_DATA, SUBMIT_INFO, ACTION_INFO = get_wts_data()
#     EMP_EMAIL, MANAGER_DATA = get_email_id()

#     SUB_TOTAL,TOTAL = get_subTotal_and_total_time()

#     idx = []
#     for employee in MANAGER_DATA:
#         tempIdx = []
#         for emp in MANAGER_DATA[employee]:
#             for each in WTS_DATA:
#                 if each[0] == emp:
#                     tempIdx.append(WTS_DATA.index(each))
#         idx.append(tempIdx)

#     temp_SUB_TOTAL_SUM = []
#     temp_TOTAL_SUM = []
#     SUB_TOTAL_SUM = []
#     TOTAL_SUM = []

#     for index in idx:
#         templist1 = []
#         templist2 = []
#         for i, each in enumerate(index):
#             templist1.append(SUB_TOTAL[i])
#             templist2.append(TOTAL[i])
#         temp_SUB_TOTAL_SUM.append(findsum(templist1))
#         temp_TOTAL_SUM.append(findsum(templist2))

#     for each in temp_SUB_TOTAL_SUM:
#         SUB_TOTAL_SUM.append(sum(each))

#     for each in temp_TOTAL_SUM:
#         TOTAL_SUM.append(sum(each))

#     data = {}
#     i = 0
#     for employee in MANAGER_DATA:
#         data[employee] = [ SUB_TOTAL_SUM[i], TOTAL_SUM[i] ]
#         i += 1

#     return data

def findsum(L):
    res = list() 
    for j in range(0, len(L[0])): 
        tmp = 0
        for i in range(0, len(L)): 
            tmp = tmp + L[i][j] 
        res.append(tmp)
    return res 


def get_modified_WTS_DATA(WTS_Data):
    for each  in WTS_Data:
        length = len(each)
        for j in range(length):
            if type(each[j]) == list:
                l=len(each[j])
                for i in range(0, l):
                    if each[j][i] == None:
                        each[j][i] = 0
    return WTS_Data

def get_subTotal_and_total_time():
    WTS_DATA,SUBMIT_INFO,ACTION_INFO = get_wts_data()
    modified_WTS_DATA = get_modified_WTS_DATA(WTS_DATA)
    SUB_TOTAL = {}
    TOTAL = {}
    proj_list = ['Miscellaneous', 'Holiday', 'Vacation'] 
    for i in range(0,len(WTS_DATA)):
        j=1
        subTotalList=[]
        totalList=[]
        while(j<len(WTS_DATA[i])):
            if WTS_DATA[i][j] not in proj_list:
                subTotalList.append(WTS_DATA[i][j+1])
                sum1 = findsum(subTotalList)
            totalList.append(WTS_DATA[i][j+1])
            sum2 = findsum(totalList)
            j=j+2

        SUB_TOTAL.update({WTS_DATA[i][0]:sum1})
        TOTAL.update({WTS_DATA[i][0]:sum2})
    return SUB_TOTAL,TOTAL
    # print('sub_total_Time:\n',sub_total)
    # print('total_Time:\n',total)

def get_all_managers():

    ALL_MANAGER_DATA = {}
    wts_wb = openpyxl.load_workbook(wts_file)
    wts_ws = wts_wb[wts_wb.sheetnames[0]]
    wts_max_row = wts_ws.max_row
    wts_max_col = wts_ws.max_column
    name_col = 0
    functional_manager_col = 0
    managers_manager_col = 0
    manager_col = 0
    # wts_max_row = 2
    for i in range(1, wts_max_col + 1):
        wts_cell = wts_ws.cell(row = 1, column = i)
        if(wts_cell.value == 'Name'): 
            name_col = i
        elif(wts_cell.value == "Functional Owner"):
            functional_manager_col = i
        elif(wts_cell.value == "Manager's Manager"):
            managers_manager_col = i
        elif(wts_cell.value == 'Reporting Manager'):
            manager_col = i
    for j in range(2, wts_max_row + 1):
        name = wts_ws.cell(row = j, column = name_col).value
        functional_manager = wts_ws.cell(row = j, column = functional_manager_col).value
        managers_manager = wts_ws.cell(row = j, column = managers_manager_col).value
        manager = wts_ws.cell(row = j, column = manager_col).value
        if(name != None ):
            if(functional_manager in ALL_MANAGER_DATA):
                if(managers_manager not in ALL_MANAGER_DATA[functional_manager] and managers_manager is not functional_manager):
                    ALL_MANAGER_DATA[functional_manager].append(managers_manager)
                # print(MANAGER_DATA)
            elif(managers_manager is not functional_manager):
                ALL_MANAGER_DATA.update({functional_manager:[managers_manager]})
            if(managers_manager in ALL_MANAGER_DATA):
                if(manager not in ALL_MANAGER_DATA[managers_manager] and manager is not managers_manager):
                    ALL_MANAGER_DATA[managers_manager].append(manager)
                # print(MANAGER_DATA)
            elif(manager is not managers_manager):
                    ALL_MANAGER_DATA.update({managers_manager:[manager]})
            # print(name,ALL_MANAGER_DATA)
    return ALL_MANAGER_DATA


def get_employees_of_manager(manager):
    MANAGERS_EMP = []
    EMP_EMAIL, MANAGER_DATA = get_email_id()
    for emp in MANAGER_DATA[manager]:
        MANAGERS_EMP.append(emp)
        if(emp in MANAGER_DATA and emp not in manager):
            MANAGERS_EMP.extend(get_employees_of_manager(emp))
    return MANAGERS_EMP
            

def get_manager_dashboard_data():
    DASHBOARD_DATA = {}
    WTS_DATA, SUBMIT_INFO, ACTION_INFO = get_wts_data()
    # EMP_EMAIL, MANAGER_DATA = get_email_id()
    SUB_TOTAL,TOTAL = get_subTotal_and_total_time()
    ALL_MANAGER_DATA = get_all_managers()

    for manager in ALL_MANAGER_DATA:
        MANAGERS_EMP=get_employees_of_manager(manager)

        SUB_TOTAL_SUM = 0

        for emp in MANAGERS_EMP:
            SUB_TOTAL_SUM += sum(SUB_TOTAL[emp] )
        DASHBOARD_DATA.update({manager:[len(MANAGERS_EMP),SUB_TOTAL_SUM]})
    return DASHBOARD_DATA



# a= get_manager_dashboard_data()
# print(a)
# get_wts_data()
# get_email_id()
# print(EMP_EMAIL)

# SUB_TOTAL,TOTAL = get_subTotal_and_total_time()
# print(SUB_TOTAL)
# print(TOTAL)
# print(SUBMIT_INFO)
# print(ACTION_INFO)
# print(MANAGER_DATA)

# data=get_manager_dashboard_data()
# print(data)
# aa=get_all_managers()
# print(aa)
