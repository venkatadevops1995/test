from vedikaweb.vedikaapi.serializers import BatchNameSerializer, VedaStudentExportsSerializer, VedaStudentSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from django.conf import settings
from openpyxl import load_workbook
from itertools import groupby
from django.db.models import Q,F
from datetime import date,datetime,timedelta

# Modles
from vedikaweb.vedikaapi.models import EmployeeMaster, PunchLogs, VedaBatch, Employee,EmployeeHierarchy, AttendanceAccessGroup, GlobalAccessFlag,VedaStudent

from vedikaweb.vedikaapi.constants import StatusCode, StudentDetailsHeadings
from vedikaweb.vedikaapi.utils import utils
from vedikaweb.vedikaapi.decorators import custom_exceptions, is_admin,jwttokenvalidator
from vedikaweb.vedikaapi.services.xlsxservice import ExcelServices

import logging
from vedikaweb.vedikaapi.services.attendance_services import AttendenceService as attendance
from django.utils import timezone
from django.db.models import When, Case, Value as V
log = logging.getLogger(__name__)

attendance_ = attendance()


class AttendanceApi(APIView):
    @jwttokenvalidator
    @custom_exceptions
    def get(self,request):
        from_date = self.request.query_params.get('from', timezone.now())
        to_date = self.request.query_params.get('to', timezone.now())
        downloadFlag = self.request.query_params.get('download', False)
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        # emp_id=auth_details['emp_id']
        emp_id = self.request.query_params.get('emp_id', None)
        if emp_id == None:
            emp_id=auth_details['emp_id']
        empid=auth_details['emp_id']
        attendance_for_all = self.request.query_params.get('all_emp', False)
        is_hr = self.request.query_params.get('is_hr', False)
        if(attendance_for_all):
            d1 = datetime.strptime(from_date , "%Y-%m-%d")
            d2 = datetime.strptime(to_date, "%Y-%m-%d")
            date_range = ((d2 - d1).days)

            if(date_range > 31):
                return Response(utils.StyleRes(False,'failure',{'date_range':'For All employee date range must me below 31 days'}), status=StatusCode.HTTP_BAD_REQUEST)


        # final_datastructure,attendance_flag,present_dates_list=attendance_.get_tt_final_datastructure(emp_id,from_date,to_date)
        final_dict = {'final_datastructure': [],'attendance_flag':[],'present_dates_list':[]}

        if not attendance_for_all:
            final_datastructure,attendance_flag,present_dates_list=attendance_.get_tt_final_datastructure(emp_id,from_date,to_date)
            final_dict['final_datastructure'].append(final_datastructure)
            final_dict['attendance_flag'].append(attendance_flag)
            final_dict['present_dates_list'].append(present_dates_list)
        else:
            if(not is_hr):
                direct_and_indirect_repoters_details = EmployeeHierarchy.objects.filter(manager_id = emp_id, emp__status=1).values('emp_id').distinct()
            else:
                direct_and_indirect_repoters_details = Employee.objects.filter(status=1).values('emp_id').distinct()
            
            # get data of user itself, if user is NOT a reporting manager of himself/herself.
            if emp_id not in [eachereporter['emp_id']for eachereporter in direct_and_indirect_repoters_details]:
                final_datastructure,attendance_flag,present_dates_list=attendance_.get_tt_final_datastructure(emp_id,from_date,to_date)
                final_dict['final_datastructure'].append(final_datastructure)
                final_dict['attendance_flag'].append(attendance_flag)
                final_dict['present_dates_list'].append(present_dates_list)
            # get data of all direct and indirect employee under user.
            for eachereporter in direct_and_indirect_repoters_details:
                emp_id = eachereporter['emp_id']
                final_datastructure,attendance_flag,present_dates_list=attendance_.get_tt_final_datastructure(emp_id,from_date,to_date)
                final_dict['final_datastructure'].append(final_datastructure)
                final_dict['attendance_flag'].append(attendance_flag)
                final_dict['present_dates_list'].append(present_dates_list)
        if(not downloadFlag):
            if attendance_for_all:
                return Response(utils.StyleRes(message='success',results={"downloadable":True}), status=StatusCode.HTTP_OK)
            else:
                return Response(utils.StyleRes(message='success',results=reversed(final_dict['final_datastructure'][0])), status=StatusCode.HTTP_OK)
        elif downloadFlag and attendance_for_all:
            emp_name = Employee.objects.filter(emp_id=empid)[0].emp_name
            basename=''+emp_name+"_Team_Attendance_"+str(from_date)+'_'+str(to_date)+'.xlsx'
            response=utils.contentTypesResponce('xl',basename)
            e=ExcelServices(response,in_memory=True,workSheetName="Attendance Report",cell_format={'font_size': 10,'font_name':'Arial','align':'left'})
            columns=['Staff No','Name','Alternative Id','Date','FirstInTime','LastOutTime','Gross Working Hours','Net Working Hours', 'Timesheet Posted Hours']
            data=[columns]

            for each_data in final_dict['final_datastructure']:
                for each in each_data:
                    data.append([each['staff_no'],each['emp_name'],each['alternative_id'],each['Date'],each['FirstInTime'],each['LastOutTime'],each['GrossWorkingHours'][:-3],each['NetWorkingHours'][:-3], each['timesheet_total_working_hours']])
                    for i,eachpunch in enumerate(each['punchdata']):
                        if('P'+str(i) not in data[0]):
                            data[0].append('P'+str(i))
                        data[-1].append(str(eachpunch['In'])+'|'+str(eachpunch['Out'])+'|'+str(eachpunch['Net']))
            e.writeExcel(data,row_start=0,long_column_list=[2])
            del e
            return response
        else:
            basename=final_datastructure[0]['emp_name']+"_Attendance_"+str(from_date)+'_'+str(to_date)+'.xlsx'
            response=utils.contentTypesResponce('xl',basename)
            e=ExcelServices(response,in_memory=True,workSheetName="Attendance Report",cell_format={'font_size': 10,'font_name':'Arial','align':'left'})
            columns=['Staff No','Name','Alternative Id','Date','FirstInTime','LastOutTime','Gross Working Hours','Net Working Hours', 'Timesheet Posted Hours']
            data=[columns]
            for each in final_datastructure:
                data.append([each['staff_no'],each['emp_name'],each['alternative_id'],each['Date'],each['FirstInTime'],each['LastOutTime'],each['GrossWorkingHours'][:-3],each['NetWorkingHours'][:-3], each['timesheet_total_working_hours']])
                for i,eachpunch in enumerate(each['punchdata']):
                    if('P'+str(i) not in data[0]):
                        data[0].append('P'+str(i))
                    data[-1].append(str(eachpunch['In'])+'|'+str(eachpunch['Out'])+'|'+str(eachpunch['Net']))
            e.writeExcel(data,row_start=0,long_column_list=[2])
            del e
            return response
        # return Response(utils.StyleRes(False,'failure',{'msg':'employee id not exists'}), status=StatusCode.HTTP_OK)

class AttendanceStatusAPI(APIView):
    @jwttokenvalidator
    @custom_exceptions
    def get(self,request):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        attendance_flag = False
        emp_id=auth_details['emp_id']
        individual_att_access_list=[]
        global_attendance_access = GlobalAccessFlag.objects.filter(status=1,access_type__iexact='ATTENDANCE')
        if(len(global_attendance_access)>0):
            att_access_grp_list = list(map(lambda x:x.emp_id,Employee.objects.filter(role_id=4,status=1)))
        else:
            att_access_grp_obj = AttendanceAccessGroup.objects.filter(status=1)
            att_access_grp_list = list(map(lambda x: x.emp_id,att_access_grp_obj))
            att_access_individ_obj = AttendanceAccessGroup.objects.filter(status=2,emp_id=emp_id)
            individual_att_access_list = list(map(lambda x: x.emp_id,att_access_individ_obj))

        emp_hierarchy_obj = EmployeeHierarchy.objects.filter(manager_id__in=att_access_grp_list,emp_id=emp_id)
        if(len(emp_hierarchy_obj)>0 or len(individual_att_access_list)>0):
            attendance_flag = True
        return Response({'attendance_flag':attendance_flag})



class VedaStudentBatchAPI(APIView):
    ''' Get All batch list'''
    @jwttokenvalidator
    @custom_exceptions
    @is_admin
    def get(self,request):
        status = request.GET.get('status','')
        if(status):
            all_batch = VedaBatch.objects.filter(status=status).order_by('-id')
        else:
            all_batch = VedaBatch.objects.order_by('-id')

        if(len(all_batch) == 0):
            return Response(utils.StyleRes(True,'No data available',{}), status=StatusCode.HTTP_NO_CONTENT)
        
        return Response(utils.StyleRes(True,'success',{'batches':all_batch.values()}), status=StatusCode.HTTP_OK)
    
    ''' Create a new batch'''
    @jwttokenvalidator
    @custom_exceptions
    @is_admin
    def post(self,request):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        data = BatchNameSerializer(data=request.data)
        
        try:
            if(data.is_valid(raise_exception=True)):
                batch_name = data["batch_name"].value.strip()
                if VedaBatch.objects.filter(batch_name = batch_name).exists():
                    return Response(utils.StyleRes(False,"Batch Name already exist",{'data':batch_name}), status= StatusCode.HTTP_NOT_ACCEPTABLE)   
                batch_details = VedaBatch.objects.create(batch_name = batch_name ,status = 1)
                return Response(utils.StyleRes(True,'batch Created successfully',{'data':data.data}), status=StatusCode.HTTP_CREATED)
        except Exception as e:
            return Response(utils.StyleRes(False,"new batch creation failed",e.args), status=409)
        

    ''' Activate or deactivate a batch'''
    @jwttokenvalidator
    @custom_exceptions
    @is_admin
    def put(self,request):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        
        data = BatchNameSerializer(data=request.data)
        try:
            if(data.is_valid(raise_exception=True)):
                batch_name = data["batch_name"].value.strip()
                status = data["status"].value

                if not VedaBatch.objects.filter(batch_name = batch_name).exists():
                    return Response(utils.StyleRes(False,"Batch Name not exist",{'data':batch_name}), status= StatusCode.HTTP_NOT_ACCEPTABLE)
                
                VedaBatch.objects.filter(batch_name = batch_name).update(status = status)
                updated_data = VedaBatch.objects.filter(batch_name = batch_name).values()[0]
                return Response(utils.StyleRes(True,'batch status updated successfully',updated_data), status=StatusCode.HTTP_OK)
        except Exception as e:
            return Response(utils.StyleRes(False,"batch updation failed",e.args), status= StatusCode.HTTP_NOT_ACCEPTABLE)

class VedaStudentAPI(APIView):
    ''' Get All student for a specific batch '''
    @jwttokenvalidator
    @custom_exceptions
    @is_admin
    def get(self,request):
        batch_name = request.GET.get('batch_name','')
        try:

            if(batch_name == '' or batch_name == None):
               return Response(utils.StyleRes(False,'batch_name is required'), status=StatusCode.HTTP_BAD_REQUEST) 
            
            if not VedaBatch.objects.filter(status=1,batch_name=batch_name).exists():
                return Response(utils.StyleRes(False,'batch_name is not exist / batch_name is inactive',{'batch_name':batch_name}), status=StatusCode.HTTP_BAD_REQUEST)
            
            batch_details = VedaBatch.objects.filter(status=1,batch_name=batch_name)
            batch_id = batch_details[0].id

            student_data = VedaStudent.objects.filter(status=1, batch_id=batch_id).order_by('-id')
            if(len(student_data) == 0):
                return Response(utils.StyleRes(True,'Data is not available',{}), status=StatusCode.HTTP_NO_CONTENT)
            final_data = student_data.values()
            return Response(utils.StyleRes(True,'success',final_data), status=StatusCode.HTTP_OK)
        except Exception as e:
            return Response(utils.StyleRes(False,"Error while try to fetch student data",e.args), status=409)
        

    ''' Create a student into the batch'''
    @jwttokenvalidator
    @custom_exceptions
    @is_admin
    def post(self,request):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        data = VedaStudentSerializer(data=request.data)
        
        try:
            if(data.is_valid(raise_exception=True)):
                batch_name = data["batch_name"].value.strip()
                student_name = data["student_name"].value.strip()
                device_id = data["device_id"].value

                if not VedaBatch.objects.filter(batch_name = batch_name,status=1).exists():
                    return Response(utils.StyleRes(False,"Batch Name is not exist / Batch is inactive",{'data':data.data}), status= StatusCode.HTTP_BAD_REQUEST)   
                
                batch_details = VedaBatch.objects.filter(batch_name = batch_name ,status = 1)
                batch_id = batch_details[0].id

                if VedaStudent.objects.filter(batch_id = batch_id,student_name =student_name,status =1).exists():
                    return Response(utils.StyleRes(False,"Student name is already exist",{'data':data.data}), status= StatusCode.HTTP_BAD_REQUEST)   
                
                new_student = VedaStudent.objects.create(batch_id = batch_id,student_name =student_name, device_id =device_id,status =1)

                return Response(utils.StyleRes(True,'New student creation successfully',{'data':data.data}), status=StatusCode.HTTP_CREATED)
        except Exception as e:
            return Response(utils.StyleRes(False,"student creation failed",e.args), status=409)
        

class ExportVedaStudentApi(APIView):


    ''' Get the excel template to upload bulk student details'''
    @jwttokenvalidator
    @custom_exceptions
    @is_admin
    def get(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request) 
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        emp_id=auth_details['emp_id']

        batch_id = kwargs['batch_id']
        print("batch_id",batch_id)

        if(not VedaBatch.objects.filter(id=batch_id,status=1).exists()):
            return Response(utils.StyleRes(False,"Batch is not exist/inactive",
        ),status=StatusCode.HTTP_BAD_REQUEST)
        batch_details = VedaBatch.objects.filter(id=batch_id).values()[0]
        batch_name = batch_details['batch_name']
        excel_file = utils.contentTypesResponce('xl',batch_name+"_"+"students_details"+".xlsx")
        excel = ExcelServices(excel_file,in_memory=True,workSheetName='students_details')
        columns = []
        for each in StudentDetailsHeadings:
            columns.append(utils.strip_value(each.value))
        excel_data= [columns]
        excel.writeExcel(excel_data,row_start=0)
        excel.terminateExcelService()
        del excel
        return excel_file
           
    @jwttokenvalidator
    @custom_exceptions
    @is_admin
    def post(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request) 
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        is_emp_admin = auth_details["is_emp_admin"]
        if(is_emp_admin == False):
            return Response(utils.StyleRes(False,"Student Bulk upload error", "user is not admin"),status=StatusCode.HTTP_UNAUTHORIZED)
        batch_id = kwargs['batch_id']
        print("batch_id",batch_id)

        if(not VedaBatch.objects.filter(id=batch_id,status=1).exists()):
            return Response(utils.StyleRes(False,"Batch is inactive /not exist",
        ),status=StatusCode.HTTP_BAD_REQUEST)

        try:
            excel_name = request.data
            excel_file_name = excel_name['file']
            uploadedfilename = str(excel_file_name).split('.')[0]
            uploadedfileext=str(excel_file_name).split('.')[-1]
            filename=uploadedfilename+"_"+utils.getUniqueId() + '.'+ uploadedfileext
            utils.createDirIfNotExists(settings.UPLOAD_ADMIN_EMAIL_ATTACHMENT_PATH)
            with open(settings.UPLOAD_ADMIN_EMAIL_ATTACHMENT_PATH+filename, 'wb') as f:
                f.write(excel_file_name.read())
            #Getting the sheet names from excel
            wb = load_workbook(excel_name['file'])
            sheet_names = wb.sheetnames
            #taking the first sheet in excel
            worksheet = wb[sheet_names[0]]
            empty_rows = []
            #getting empty rows deleting from worksheet
            for idx, row in enumerate(worksheet.iter_rows(max_col=worksheet.max_column), start=1):
                empty = not any((utils.strip_value(cell.value) for cell in row))
                if empty:
                    empty_rows.append(idx)
            for row_idx in reversed(empty_rows):
                worksheet.delete_rows(row_idx, 1)

            #case insensitive for the headings
            for each in worksheet[1]:
                if each.value != None and type(each.value) != int:
                    each.value = each.value.lower()
                else:
                    each.value
            #needed columns form excel
            need_columns =[]
            for each in StudentDetailsHeadings:
                need_columns.append(utils.strip_value(each.value))

            #taking all excel columns
            excel_columns=[]
            for cell in worksheet[1]:
                if utils.strip_value(cell.value) != None:
                    excel_columns.append(cell.value)
                else:
                    excel_columns.append(cell.value)
            log.debug('excel_columns: '+",".join(map(str,excel_columns)))

            row_count = worksheet.max_row #max_row count from excel
            #checking all needed columns are present in excel
            main_list = list(set(need_columns) - set(excel_columns))
            #is headings unique
            is_column_headings_unique  = False
            is_column_headings_unique = (len(set(excel_columns)) == len(excel_columns))

            if(len(main_list)>0 or is_column_headings_unique==False):
                return Response(utils.StyleRes(False,"Student Bulk upload error", {'missing_columns':main_list,'unique_columns':is_column_headings_unique}
            ),status=StatusCode.HTTP_BAD_REQUEST)
            #taking all the student names from excel Student Name  row columns and validating all the student names
            valid_studentnames = []
            valid_device_ids = []
            invalid_studentnames =[]
            invalid_device_ids=[]
            duplicated_studentnames = []
            duplicated_device_ids = []
            student_data = []
            uploaded_student_list = []
            if len(main_list) == 0:
                elem = {k: i for i, k in enumerate(excel_columns)}
                output = list(map(elem.get, need_columns))
                row_count = worksheet.max_row
                for i in range(row_count + 1):
                    if i > 1:
                        column_details = [cell.value for cell in worksheet[i]]
                        # print('email',column_details)
                        student_name = utils.strip_value(column_details[output[0]])
                        device_id = utils.strip_value(column_details[output[1]])
                        
                        if(VedaStudent.objects.filter(batch_id=batch_id,student_name=student_name,status=1).exists()):
                            print("duplicate student ...",student_name)
                            err = []
                            err_obj={}
                            err_obj['student_name']= '{} ,student name is already exist'.format(student_name)
                            err.append(err_obj)                           
                            return Response(utils.StyleRes(False,"Student Bulk upload error", err ),status=StatusCode.HTTP_BAD_REQUEST)
                        
                        if(student_name not in valid_studentnames):
                            valid_studentnames.append(student_name)
                        else:
                            duplicated_studentnames.append(student_name)

                        if(device_id not in valid_device_ids):
                            valid_device_ids.append(device_id)
                        else:
                            duplicated_device_ids.append(device_id)
                        
                        if(student_name == ''):
                            invalid_studentnames.append(student_name)

                        if(device_id == ''):
                            invalid_device_ids.append(device_id)

                        student_data.append({'student_name':student_name,'device_id':device_id,'batch':batch_id,'status':1})
                        uploaded_student_list.append(student_name)
                res = {'duplicated_studentnames':duplicated_studentnames,'invalid_studentnames':invalid_studentnames,
                    'duplicated_device_ids':duplicated_device_ids, 'invalid_device_ids':invalid_device_ids,
                }

                if(len(invalid_studentnames)>0 or len(duplicated_studentnames)>0 or len(invalid_device_ids)>0 or len(duplicated_device_ids)>0):
                    return Response(utils.StyleRes(False,"Student Bulk upload error", res),status=StatusCode.HTTP_BAD_REQUEST)
                
                serial_student_data = VedaStudentExportsSerializer(data=student_data,many=True)
                if(serial_student_data.is_valid()):
                    serial_student_data.save()
                    res.update({'uploaded_student_list':uploaded_student_list})
                    return Response(utils.StyleRes(True,"Student data uploaded successfully", res),status=StatusCode.HTTP_CREATED)
                else:
                    return Response(utils.StyleRes(False,"Student Bulk upload error", serial_student_data.errors),status=StatusCode.HTTP_BAD_REQUEST)
        except Exception as e:
            print("e.........", e)
            return Response(utils.StyleRes(False,"Student Bulk upload error",e.args), status=409)


class VedaStudentAttendanceApi(APIView):
    @jwttokenvalidator
    @custom_exceptions
    @is_admin
    def get(self,request):
        batch_name = request.GET.get('batch_name','')
        from_date =  request.GET.get('start_date','')
        to_date = request.GET.get('end_date','')

        if(batch_name == '' or from_date == '' or to_date == ''):
            return Response(utils.StyleRes(False,'failure',{'msg':'batch_name, start date and end date  is required '}), status=StatusCode.HTTP_NOT_ACCEPTABLE)
        
        batchDetails = VedaBatch.objects.filter(status=1,batch_name=batch_name)
 
        if(len(batchDetails) == 0):
            return Response(utils.StyleRes(False,'No content',{'msg':'no content '}), status=StatusCode.HTTP_NO_CONTENT)
        
        batch_id = batchDetails[0].id
        all_student_data = VedaStudent.objects.filter(batch_id=batch_id,status=1).values()
        deviceIDList = []
        final_dict = {'final_datastructure': []}

        for student in all_student_data:
            deviceIDList.append(student['device_id'])
        for deviceId in deviceIDList:
            final_datastructure = attendance_.get_student_final_datastructure(deviceId,from_date,to_date)
            final_dict['final_datastructure'].append(final_datastructure)
        columns=['DeviceId','Date','FirstInTime','LastOutTime','Gross Working Hours','Net Working Hours','Attendance','isHoliday']

        file_name=batchDetails[0].batch_name+"_Student_Attendance_"+str(from_date)+'_'+str(to_date)+'.xlsx'
        response=utils.contentTypesResponce('xl',file_name)
        e=ExcelServices(response,in_memory=True,workSheetName="Student Attendance Report",cell_format={'font_size': 10,'font_name':'Arial','align':'left'})
        data=[columns]
        
        for each_data in final_dict['final_datastructure']:
            for each in each_data:
                data.append([each['deviceId'],each['Date'],each['FirstInTime'],each['LastOutTime'],each['GrossWorkingHours'][:-3],each['NetWorkingHours'][:-3],each['attendance'], 'Yes'  if each['isHoliday'] == True else 'No'])
        e.writeExcel(data,row_start=0,long_column_list=[2])
        del e
        return response
    
class AttendanceByAltId(APIView):
    @jwttokenvalidator
    @custom_exceptions
    # @is_admin
    def get(self, request):
        from_date =  request.GET.get('start_date','')
        to_date = request.GET.get('end_date','')
        format = "%Y-%m-%d"
        if(from_date == '' or to_date == ''):
            return Response(utils.StyleRes(False,'failure',{'msg':'start date and end date  is required '}), status=StatusCode.HTTP_NOT_ACCEPTABLE)
        all_emp_data = Employee.objects.filter(status=1)
        ignorePunchLog = settings.IGNOROR_PUNCH_DEVICES 
        final_dict = {'final_datastructure': []}
        if(type(from_date)==str):
            from_date_with_time = datetime.strptime(from_date,format)
        if(type(to_date)==str):
            to_date_with_time = datetime.strptime(to_date,format)
        to_date_with_time = datetime.combine(to_date_with_time, datetime.min.time())+timedelta(hours=23,minutes=59,seconds=59)
        for each_emp in all_emp_data:
            staff_no = each_emp.staff_no
            emp_name = each_emp.emp_name
            emp_master_data = EmployeeMaster.objects.using('attendance').filter( Q(EmpId=staff_no) & Q(AmdId__gt = 0) & Q(DeviceId__gt = 0))
            
            each_emp_data = {'data':'','emp_name':emp_name}
            if(len(emp_master_data) > 0):
                deviceId = emp_master_data.last().DeviceId
                amdId = emp_master_data.last().AmdId
                each_emp_punch_data = PunchLogs.objects.using('attendance').filter( Q(DeviceID=deviceId) & Q(LogDate__gte=from_date_with_time) & Q(LogDate__lte=to_date_with_time)  & (~Q(SerialNo__in=ignorePunchLog))).order_by('LogDate')
                each_emp_data['data'] = each_emp_punch_data.values()
                each_emp_data['amdId'] = amdId
                
            final_dict['final_datastructure'].append(each_emp_data)

        columns=[]

        file_name="Attendance_By_Alt_Id_"+str(from_date)+'_'+str(to_date)+'.xlsx'
        response=utils.contentTypesResponce('xl',file_name)
        e=ExcelServices(response,in_memory=True,workSheetName="Attendance By Alt.Id Report",cell_format={'font_size': 10,'font_name':'Arial','align':'left'})
        data=[columns]
        data.append(['SOCTRONICS TECHNOLOGIES PVT. LTD.'])
        data.append([])
        data.append([])
        data.append(['From Date: {} To Date: {}'.format(from_date, to_date)])
        data.append([])
        data.append([])
        data.append(['Badge ID','EmpName','DateTime','In/Out','ReaderName'])


        for each_data in final_dict['final_datastructure']:
            for each in each_data['data']:
                logDate = each['LogDate'].strftime('%Y-%m-%d %H:%M:%S')
                data.append([each_data['amdId'],each_data['emp_name'],logDate, 'I' if each['Direction'] == 'In' else 'O' ,each['Source']])

        e.writeExcel(data,row_start=0,long_column_list=[2])
        del e
        return response

        