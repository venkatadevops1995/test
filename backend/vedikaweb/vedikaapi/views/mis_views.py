from vedikaweb.vedikaapi.services.email_service import email_service
from rest_framework.views import APIView
from rest_framework.response import Response
from openpyxl import load_workbook
import time 
import datetime 
from datetime import date
# Modles
from vedikaweb.vedikaapi.models import Employee,EmployeeProject,Project,MisInfo,Project,Location,EmployeeHierarchy, GlobalAccessFlag, WelcomeEmailNotification, EmailAccessGroup,ServiceAccount,Category, NewHireLeaveConfig, LeaveBalance, LeaveAccessGroup,EmployeeProjectTimeTracker

# Serialisers
from vedikaweb.vedikaapi.serializers import EmployeeProfileSerializer,MisDonloadDisableWithDate

from vedikaweb.vedikaapi.constants import StatusCode, ExcelHeadings, DefaultProjects, MailConfigurations, GenderChoices, MaritalStatuses
from vedikaweb.vedikaapi.utils import utils
from django.conf import settings
from vedikaweb.vedikaapi.decorators import custom_exceptions,jwttokenvalidator
from django.db.models import Q,F,CharField,Value as V , Sum

from openpyxl.utils.cell import get_column_letter
from django.core.validators import  validate_email
from vedikaweb.vedikaapi.services.xlsxservice import ExcelServices
import traceback, json
from datetime import datetime, timedelta
import logging
from vedikaweb.vedikaapi.services.attendance_services import AttendenceService as attendance
from hashlib import md5
from django.core.mail import send_mail
from django.template.loader import get_template
from django.db.models import When, Case 
from django.db.models.functions import Coalesce
log = logging.getLogger(__name__)
import operator
attendance_ = attendance()


class MisUpload(APIView):
    @custom_exceptions
    def post(self, request,*args,**kwargs):
        start_time = time.time()
        #getting the excel from request
        # print("=========",request.data)
        excel_name = request.data
        service_details = ServiceAccount.objects.filter(api_user='mis_apikey')
        if(len(service_details)==0):
            log.error("failed to get api_key from database")
            return Response(utils.StyleRes(False,'failed to get api_key from database',{}), status=StatusCode.HTTP_UNPROCESSABLE_ENTITY)
        else:
            password = service_details[len(service_details)-1].password
            hashpass = md5(excel_name['api_key'].encode('utf8')).hexdigest()
            if(password!=hashpass):
                log.error("api_key not matched")
                return Response(utils.StyleRes(False,'api_key not matched',{}), status=StatusCode.HTTP_UNPROCESSABLE_ENTITY)

        print('excel_name',excel_name['file'])
        #getting excel and updating the excel name with cuttent time stamp and saving excel in directory
        excel_file_name = excel_name['file']
        uploadedfilename = str(excel_file_name).split('.')[0]
        uploadedfileext=str(excel_file_name).split('.')[-1]
        filename=uploadedfilename+"_"+utils.getUniqueId() + '.'+ uploadedfileext
        utils.createDirIfNotExists(settings.UPLOAD_PATH)
        dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(settings.UPLOAD_PATH+filename, 'wb') as f:
            f.write(excel_file_name.read())
        #Getting the sheet names from excel
        wb = load_workbook(excel_name['file'])
        sheet_names = wb.sheetnames
        #taking the first sheet in excel
        worksheet = wb[sheet_names[0]]
        # print('worksheet',len(worksheet[1]))
        # print('max_column_length',worksheet.max_row)
        empty_rows = []
        #getting empty rows deleting from worksheet
        for idx, row in enumerate(worksheet.iter_rows(max_col=worksheet.max_column), start=1):
            empty = not any((utils.strip_value(cell.value) for cell in row))
            if empty:
                empty_rows.append(idx)
        for row_idx in reversed(empty_rows):
            worksheet.delete_rows(row_idx, 1)

        #case insensitive for the headings
        for each in worksheet[1]:
            if each.value != None and type(each.value) != int:
                each.value = each.value.lower()
            else:
                each.value

        #needed columns form excel
        need_columns =[]
        for each in ExcelHeadings:
            need_columns.append(utils.strip_value(each.value))

        #taking all excel columns
        excel_columns=[]
        for cell in worksheet[1]:
            if utils.strip_value(cell.value) != None:
                excel_columns.append(cell.value)
            else:
                 excel_columns.append(cell.value)
        print('excel_columns ',excel_columns)
            # else:
            #     if (utils.strip_value(cell.value) != None):
            #         excel_columns.append(cell.value)

        # print('ExcelHeadings.company',excel_columns)
        row_count = worksheet.max_row #max_row count from excel
        #checking all needed columns are present in excel
        main_list = list(set(need_columns) - set(excel_columns))
        #is headings unique
        is_column_headings_unique  = False
        is_column_headings_unique = (set(excel_columns) == len(excel_columns))
        print(main_list)
        #taking all the emails from excel Email row columns and validating all the emails
        invalid_emails = []
        valid_emails = []
        valid_empnames = []
        valid_staffnos = []
        all_managers_list = []
        empty_empnames = []
        empty_staffnos = []

        all_date_of_joinings = []
        empty_date_of_joinings = []
        invalid_date_of_joinings = []

        all_locations = []
        empty_locations = []
        invalid_locations = []

        all_marital_statuses = []
        empty_marital_statuses = []
        invalid_marital_statuses = []

        all_genders = []
        empty_genders = []
        invalid_genders = []

        all_categories = []
        empty_categories = []
        invalid_categories = []
        




        if len(main_list) == 0:
            print("Yes, list1 contains all elements in list2")
            #checking same project names for employee
            same_projects = []

            elem = {k: i for i, k in enumerate(excel_columns)}
            output = list(map(elem.get, need_columns))
            row_count = worksheet.max_row
            for i in range(row_count + 1):
                if i > 1:
                    column_details = [utils.strip_value(cell.value) if not isinstance(cell.value,datetime) else cell.value for cell in worksheet[i]]
                    # print('email',column_details)
                    emp_name = utils.strip_value(column_details[output[3]])
                    project1 = utils.strip_value(column_details[output[4]])
                    project2 = utils.strip_value(column_details[output[5]])
                    project3 = utils.strip_value(column_details[output[6]])
                    emp_projects = []
                    if project1 != None and project1 != 0 and project1 != '#N/A':
                        emp_projects.append(project1)
                    if project2 != None and project2 != 0 and project2 != '#N/A':
                        emp_projects.append(project2)
                    if project3 != None and project3 != 0 and project3 != '#N/A':
                        emp_projects.append(project3)
                    is_projects_unique = False
                    is_projects_unique = len(set(emp_projects)) == len(emp_projects)
                    if is_projects_unique == False:
                        same_projects.append({'Name':emp_name,'projects':str(project1) + ',' + str(project2) + ',' + str(project3)})

            #taking all emails and checking email format
            for cell in worksheet[1]:
                if (utils.strip_value(cell.value) == ExcelHeadings.Email.value):

                    row_no = get_column_letter(cell.column)
                    # print(worksheet[row_no])
                    for column in worksheet[row_no]:
                        if column.value != ExcelHeadings.Email.value:
                            # print('column',column.value)
                            try:
                                validate_email(utils.strip_value(column.value))
                            except:
                                log.error(traceback.format_exc())
                                invalid_emails.append(utils.strip_value(column.value))
                                pass
                            else:
                                valid_emails.append(utils.strip_value(column.value))
                #taking all empnames
                if (utils.strip_value(cell.value) == ExcelHeadings.Emp_name.value):
                    row_no = get_column_letter(cell.column)
                    for column in worksheet[row_no]:
                        if utils.strip_value(column.value) != ExcelHeadings.Emp_name.value:
                            if (utils.strip_value(column.value) != None) and (utils.strip_value(column.value) != 0) and (utils.strip_value(column.value) != '#N/A'):
                                valid_empnames.append(utils.strip_value(column.value))
                            else:
                                empty_empnames.append(utils.strip_value(column.value))

                #taking all staffnos
                if (utils.strip_value(cell.value) == ExcelHeadings.Empid.value.lower()):
                    row_no = get_column_letter(cell.column)
                    for column in worksheet[row_no]:
                        if utils.strip_value(column.value) != ExcelHeadings.Empid.value:
                            if (utils.strip_value(column.value) != None) and (utils.strip_value(column.value) != 0) and (utils.strip_value(column.value) != '#N/A'):
                                valid_staffnos.append(utils.strip_value(column.value))
                            else:
                                empty_staffnos.append(utils.strip_value(column.value))
                                
                # print('empty_staffnos',len(empty_staffnos))
                #taking all Magagers list
                if (utils.strip_value(cell.value) == ExcelHeadings.Rep_manager.value):
                    row_no = get_column_letter(cell.column)
                    for column in worksheet[row_no]:
                        if utils.strip_value(column.value) != ExcelHeadings.Rep_manager.value:
                            if (utils.strip_value(column.value) != None) and (utils.strip_value(column.value) != 0) and (utils.strip_value(column.value) != '#N/A'):
                                all_managers_list.append(utils.strip_value(column.value))
                if (utils.strip_value(cell.value) == ExcelHeadings.Manager_manager.value):
                    row_no = get_column_letter(cell.column)
                    for column in worksheet[row_no]:
                        if utils.strip_value(column.value) != ExcelHeadings.Manager_manager.value:
                            if (utils.strip_value(column.value) != None) and (utils.strip_value(column.value) != 0) and (utils.strip_value(column.value) != '#N/A'):
                                all_managers_list.append(utils.strip_value(column.value))
                if (utils.strip_value(cell.value) == ExcelHeadings.Fun_owner.value):
                    row_no = get_column_letter(cell.column)
                    for column in worksheet[row_no]:
                        if utils.strip_value(column.value) != ExcelHeadings.Fun_owner.value:
                            if (utils.strip_value(column.value) != None) and (utils.strip_value(column.value) != 0) and (utils.strip_value(column.value) != '#N/A'):
                                all_managers_list.append(utils.strip_value(column.value))
                            # print("bad email, details:", cell.value)

                if (utils.strip_value(cell.value) == ExcelHeadings.Location.value.lower()):
                    row_no = get_column_letter(cell.column)
                    for column in worksheet[row_no]:
                        if utils.strip_value(column.value) != ExcelHeadings.Location.value:
                            if (utils.strip_value(column.value) != None) and (utils.strip_value(column.value) != 0) and (utils.strip_value(column.value) != '#N/A'):
                                all_locations.append(utils.strip_value(column.value))
                            else:
                                empty_locations.append(utils.strip_value(column.value))
                if (utils.strip_value(cell.value) == ExcelHeadings.Gender.value.lower()):
                    row_no = get_column_letter(cell.column)
                    for column in worksheet[row_no]:
                        if utils.strip_value(column.value) != ExcelHeadings.Gender.value:
                            if (utils.strip_value(column.value) != None) and (utils.strip_value(column.value) != 0) and (utils.strip_value(column.value) != '#N/A'):
                                all_genders.append(utils.strip_value(column.value))
                            else:
                                empty_genders.append(utils.strip_value(column.value))
                if (utils.strip_value(cell.value) == ExcelHeadings.Marital_status.value.lower()):
                    row_no = get_column_letter(cell.column)
                    for column in worksheet[row_no]:
                        if utils.strip_value(column.value) != ExcelHeadings.Marital_status.value:
                            if (utils.strip_value(column.value) != None) and (utils.strip_value(column.value) != 0) and (utils.strip_value(column.value) != '#N/A'):
                                all_marital_statuses.append(utils.strip_value(column.value))
                            else:
                                empty_marital_statuses.append(utils.strip_value(column.value))
                if (utils.strip_value(cell.value) == ExcelHeadings.Date_of_joining.value.lower()):
                    row_no = get_column_letter(cell.column)
                    for column in worksheet[row_no]:
                        if column.value != ExcelHeadings.Date_of_joining.value or (not  isinstance(column.value,datetime) and utils.strip_value(cell.value) == ExcelHeadings.Marital_status.value.lower()):
                            if ((column.value != None) and (column.value != 0) and (column.value != '#N/A') and  isinstance(column.value,datetime)):
                                all_date_of_joinings.append(column.value)
                            else:
                                empty_date_of_joinings.append(column.value)

                if (utils.strip_value(cell.value) == ExcelHeadings.Category.value.lower()):
                    row_no = get_column_letter(cell.column)
                    for column in worksheet[row_no]:
                        if utils.strip_value(column.value) != ExcelHeadings.Category.value:
                            if (utils.strip_value(column.value) != None) and (utils.strip_value(column.value) != 0) and (utils.strip_value(column.value) != '#N/A'):
                                all_categories.append(utils.strip_value(column.value))
                            else:
                                empty_categories.append(utils.strip_value(column.value))

                

            existing_locations = list(Location.objects.filter(status=1))
            empty_locations.extend(list(filter(lambda x: not any( x.lower()==each_loc.name.lower() for each_loc in existing_locations) ,all_locations)))

            existing_genders =  [ g.name for g in GenderChoices]
            empty_genders.extend(list(filter(lambda x: not any( x.upper()==each_gen for each_gen in existing_genders) ,all_genders)))

            existing_marital_statuses= [ m.name for m in  MaritalStatuses]
            empty_marital_statuses.extend(list(filter(lambda x: not any( x.upper()==each_mrrg_st for each_mrrg_st in existing_marital_statuses) ,all_marital_statuses)))
            
            existing_categories = list(Category.objects.filter(status=1))
            empty_categories.extend(list(filter(lambda x: not any( x.lower()==each_cat.name.lower() for each_cat in existing_categories) ,all_categories)))



            #checking the dupicate emails in the list
            is_emails_unique = False
            is_emails_unique = len(set(valid_emails)) == len(valid_emails)

            #checking the dupicate empnames in the list
            is_empname_unique = False
            is_empname_unique = len(set(valid_empnames)) == len(valid_empnames)
            #checking the dupicate staffnos in the list
            is_empid_unique = False
            is_empid_unique = len(set(valid_staffnos)) == len(valid_staffnos)
            #chceking all managers name present in the emp name list
            missing_man_list = list(set(all_managers_list) - set(valid_empnames))
            #Checking all conditions are satisfied or not

            if (len(invalid_emails) == 0) and is_emails_unique and is_empname_unique and is_empid_unique and len(missing_man_list) == 0 \
            and len(empty_empnames) == 0 and len(empty_staffnos) == 0 and len(same_projects) == 0 and len(empty_date_of_joinings) == 0 and len(empty_locations) == 0 and len(empty_genders) == 0 and len(empty_marital_statuses) == 0 and len(empty_categories) ==0:
                print("--- %s seconds ---checking" % (time.time() - start_time))

                global_leave_access = GlobalAccessFlag.objects.filter(status=1,access_type__iexact='LEAVE')
                leave_access_grp_list = []
                if(len(global_leave_access)>0):
                    leave_access_grp_list = list(map(lambda x:x.emp_id,Employee.objects.filter(role_id=4,status=1)))
                else:
                    leave_access_grp_obj = LeaveAccessGroup.objects.filter(status=1)
                    leave_access_grp_list = list(map(lambda x: x.emp_id,leave_access_grp_obj))



                #Adding employee in the database if invalid emails and duplicate data is null
                #getting excel columns and taking the need columns row
                added_emp_info = []
                elem = {k: i for i, k in enumerate(excel_columns)}
                output = list(map(elem.get, need_columns))

                row_count = worksheet.max_row #max_row count from excel
                #taking the all rows which we need to add in employee table
                '''
                    By default inserting all employees in databse
                '''
                all_emp_profile_data = []
                for i in range(row_count + 1):
                    if i > 1:
                        #getting one whole columns details
                        column_details = [utils.strip_value(cell.value) if not isinstance(cell.value,datetime) else cell.value for cell in worksheet[i]]
                        # print('email',column_details)
                        email = utils.strip_value(column_details[output[0]])
                        compnay = utils.strip_value(column_details[output[1]])
                        staff_no = utils.strip_value(column_details[output[2]])
                        emp_name = utils.strip_value(column_details[output[3]])
                        date_of_joining = column_details[output[10]]
                        fun_owner = utils.strip_value(column_details[output[7]])

                        location =  list(filter(lambda x:  utils.strip_value(column_details[output[11]]).lower()==x.name.lower(),existing_locations))[0].id
                        marital_status = MaritalStatuses[utils.strip_value(column_details[output[12]]).upper()].value
                        gender =  GenderChoices[utils.strip_value(column_details[output[13]]).upper()].value 
                        category =  list(filter(lambda x:  utils.strip_value(column_details[output[14]]).lower()==x.name.lower(),existing_categories))[0].id
                        emp_profile_data={'gender_id':gender,'location':location,'date_of_join':datetime.strftime(date_of_joining,'%Y-%m-%d'),'category':category, 'is_married':marital_status}


                        #checking the employee with email, if not avalibale inserting the Employee
                        email_check = Employee.objects.filter(email = email)
                        if len(email_check) == 0:
                            emp = Employee(email = email,password = '',emp_name = emp_name,company = compnay, staff_no = staff_no, role_id = 1,status = 1)
                            emp.save()
                            added_emp_info.append(email)
                            log.info("new employee added "+str(emp.emp_id))
                            emp_profile_data.update({'emp':emp.emp_id})
                            

                            if((fun_owner !=0) and (fun_owner != None) and (fun_owner != '#N/A') and  (len(global_leave_access)>0 or (fun_owner in leave_access_grp_list) )):
                                doj =date_of_joining
                                today = datetime.now()
                                month_diff = (today.year - doj.year)*12 + (today.month - doj.month)
                                days = doj.strftime("%d")

                                highest_credit = 0
                                new_hire_config = NewHireLeaveConfig.objects.filter(Q(category=category)&Q(time_period__start_date=1))
                                if(len(new_hire_config)>0):
                                    highest_credit = new_hire_config[0].round_off_leave_credit
                                leave_credits=NewHireLeaveConfig.objects.filter(Q(category=category)&Q(time_period__start_date__lte=days)&Q(time_period__end_date__gte=days)&Q(status=1))
                                if(len(leave_credits)==0):
                                    leave_credits = 0
                                else:
                                    leave_credits = leave_credits[0].round_off_leave_credit

                                total_leave_credit=leave_credits+ (highest_credit * month_diff)
                                
                                LeaveBalance(emp_id=emp.emp_id,year=today.year,month=today.month,leave_credits=total_leave_credit,status=1).save()
                                print("total leave credited {} for emp_id {}".format(total_leave_credit,emp.emp_id))
                                log.info("total leave credited {} for emp_id {}".format(total_leave_credit,emp.emp_id))
                                # print('email',email)
                            else:
                                log.info("Leave balance is not added for emp_id {} as Leave access is not enabled".format(emp.emp_id))
                        else:
                            Employee.objects.filter(email = email).update(emp_name = emp_name,company = compnay, staff_no = staff_no,role_id = 1)
                            emp_profile_data.update({'emp':email_check[0].emp_id})
                        all_emp_profile_data.append(emp_profile_data)
                serial_emp_profile_data=EmployeeProfileSerializer(data=all_emp_profile_data,many=True)
                if(serial_emp_profile_data.is_valid()):
                    serial_emp_profile_data.save()
                    # print(list(serial_emp_profile_data.data)[0].get('emp'))
                    log.info("employee_profile updated for emp_id: "+','.join(map(str,map(lambda x:x.get('emp'),serial_emp_profile_data.data))))
                else:
                    log.error(serial_emp_profile_data.errors)

                #checking all the emails with databse employee emails, which emails not present in the excel updationg the status to 0
                deleted_emp_info = []
                emp_details = Employee.objects.filter(status = 1)
                database_emp_emails = []
                for each in emp_details:
                    database_emp_emails.append(each.email)
                missed_emails = list(set(database_emp_emails) - set(valid_emails))
                deleted_emp_info = missed_emails
                if len(missed_emails):
                    for email in missed_emails:
                        emp_obj = Employee.objects.filter(email = email).update(status=0)
                #checking old emails are present in the excel then updating that employee status to 1
                emp_details = Employee.objects.filter(status = 0)
                emp_emails = []
                for each in emp_details:
                    emp_emails.append(each.email)
                old_emails = list(set(emp_emails) & set(valid_emails))
                if len(old_emails) > 0:
                    for email in old_emails:
                        emp_obj = Employee.objects.filter(email = email).update(status=1)
                log.info("--- %s seconds ---empupdate" % (time.time() - start_time))
                #projects checking and Updation
                '''
                    Inserting and updating  projects according to excel
                '''
                actual_projects = []
                #taking all the project list to one list and getting unique projects
                projects = [ExcelHeadings.Project1.value,ExcelHeadings.Project2.value,ExcelHeadings.Project3.value]
                for project in projects:
                    for cell in worksheet[1]:
                        if (utils.strip_value(cell.value) == project):
                            cell = get_column_letter(cell.column)
                            for cell in worksheet[cell]:
                                if (utils.strip_value(cell.value) != project) and (cell.value != None) and (cell.value != 0):
                                    actual_projects.append(utils.strip_value(cell.value))
                #getting unique projects and saving in project table if project is not present
                unique_projects = list(set(actual_projects))

                for project in unique_projects:
                    proj_data = Project.objects.filter(name = project)
                    if len(proj_data) == 0:
                        proj_create = Project(name = project,code='',status = 1)
                        proj_create.save()
                database_projects_details = Project.objects.filter(status = 1)
                database_projects =[]
                for proj in database_projects_details:
                    database_projects.append(proj.name)
                #checing database projects with unique projects, updating the status to 0 to projects which projects are not present in the excel
                missed_projects = list(set(database_projects) - set(unique_projects))
                if len(missed_projects) > 0:
                    for proj in missed_projects:
                        proj_obj = Project.objects.filter(name = proj).update(status=0)
                print("--- %s seconds ---project" % (time.time() - start_time))
                #if old projects are preent in th excel then updating that project status to 1
                proj_details = Project.objects.filter(status = 0)
                proj_names = []
                for each in proj_details:
                    proj_names.append(each.name)
                old_projects = list(set(proj_names) & set(unique_projects))
                if len(old_projects) > 0:
                    for proj in old_projects:
                        proj_obj = Project.objects.filter(name = proj).update(status=1)
                rep_manager = []
                #role Update
                #getting the report manager row and updating the role in employee table accordingly
                for cell in worksheet[1]:
                    if (utils.strip_value(cell.value) == ExcelHeadings.Rep_manager.value):
                        cell = get_column_letter(cell.column)
                        for cell in worksheet[cell]:
                            if (utils.strip_value(cell.value) != ExcelHeadings.Rep_manager.value) and (cell.value != None) and (cell.value != 0):
                                rep_manager.append(utils.strip_value(cell.value))
                unique_rep_list = list(set(rep_manager))
                # print('unique_rep_list',unique_rep_list)
                for emp_name in unique_rep_list:
                    Employee.objects.filter(emp_name = emp_name).update(role_id = 2)
                #getting the manager's manager row and updating the role
                man_manager = []
                for cell in worksheet[1]:
                    if (utils.strip_value(cell.value) == ExcelHeadings.Manager_manager.value):
                        cell = get_column_letter(cell.column)
                        for cell in worksheet[cell]:
                            if (utils.strip_value(cell.value) != ExcelHeadings.Manager_manager.value) and (cell.value != None) and (cell.value != 0):
                                man_manager.append(utils.strip_value(cell.value))
                unique_man_list = list(set(man_manager))
                for emp_name in unique_man_list:
                    Employee.objects.filter(emp_name = emp_name).update(role_id = 3)
                #getting the functional ownwer row and updating the role
                fun_owner = []
                for cell in worksheet[1]:
                    if (utils.strip_value(cell.value) == ExcelHeadings.Fun_owner.value):
                        cell = get_column_letter(cell.column)
                        for cell in worksheet[cell]:
                            if (utils.strip_value(cell.value) != ExcelHeadings.Fun_owner.value) and (cell.value != None) and (cell.value != 0):
                                fun_owner.append(utils.strip_value(cell.value))
                unique_fun_owner = list(set(fun_owner))
                for emp_name in unique_fun_owner:
                    Employee.objects.filter(emp_name = emp_name).update(role_id = 4)
                log.info("--- %s seconds ---role" % (time.time() - start_time))

                '''
                    Inserting and updating  Employee Hierarchy according to excel
                '''
                output = list(map(elem.get, need_columns))
                row_count = worksheet.max_row
                #getting the all managers for employee
                for i in range(row_count + 1):
                    if i > 1:
                        column_details = [utils.strip_value(cell.value) if not isinstance(cell.value,datetime) else cell.value for cell in worksheet[i]]
                        # print('email',column_details)
                        email = utils.strip_value(column_details[output[0]])
                        fun_owner = utils.strip_value(column_details[output[7]])
                        man_manager = utils.strip_value(column_details[output[8]])
                        rep_manager = utils.strip_value(column_details[output[9]])
                        emp_det = Employee.objects.get(email = email)
                        emp_id = emp_det.emp_id
                        #checking rep_manager column is empty, 0 and #N/A
                        if (rep_manager !=0) and (rep_manager != None) and (rep_manager != '#N/A'):
                            #checking the emp_id and project_id already present in the db if not inserting with priority is 1
                            #print('emp_name',utils.strip_value(rep_manager))
                            utils.emp_hierarchyupdate(emp_id,rep_manager,1)

                        else:
                            #if rep_manager is null updating status to zero for previuos prev_rep_manager
                            emp_hierarchy = EmployeeHierarchy.objects.filter(emp_id = emp_id,priority = 1,status = 1).update(status = 0)

                        #doing same for the manager's manager
                        if (man_manager !=0) and (man_manager != None) and (man_manager != '#N/A'):
                            utils.emp_hierarchyupdate(emp_id,man_manager,2)

                        else:
                            emp_hierarchy = EmployeeHierarchy.objects.filter(emp_id = emp_id,priority = 2,status = 1).update(status = 0)

                        #doing same for the functional owner
                        if (fun_owner !=0) and (fun_owner != None) and (fun_owner != '#N/A'):
                            utils.emp_hierarchyupdate(emp_id,fun_owner,3)

                        else:
                            emp_hierarchy = EmployeeHierarchy.objects.filter(emp_id = emp_id,priority = 3,status = 1).update(status = 0)

                log.info("--- %s seconds ---emphie" % (time.time() - start_time))
                '''
                    Inserting and updationg  emplyee_project mapping according to excel
                '''
                #taking the email and project details of an employee
                output = list(map(elem.get, need_columns))
                row_count = worksheet.max_row
                for i in range(row_count + 1):
                    if i > 1:
                        column_details = [utils.strip_value(cell.value) if not isinstance(cell.value,datetime) else cell.value  for cell in worksheet[i]]
                        # print('email',column_details)
                        email = utils.strip_value(column_details[output[0]])
                        project1 = utils.strip_value(column_details[output[4]])
                        project2 = utils.strip_value(column_details[output[5]])
                        project3 = utils.strip_value(column_details[output[6]])
                        emp_det = Employee.objects.get(email = email)
                        emp_id = emp_det.emp_id
                        #updating all priority and status to zero for the employee
                        EmployeeProject.objects.filter(emp_id = emp_id).update(priority = 0,status = 0)
                        #checking project column 0, empty and #N/A
                        if (project1 != 0) and (project1 != None) and (project1 != '#N/A'):
                            utils.emp_project_update(emp_id,project1,1)

                        else:
                            #if project value in none or empty updating the previous priority 1 project status to 0
                            EmployeeProject.objects.filter(emp_id = emp_id,priority = 1,status = 1).update(status = 0)

                        #project2 updation as same of project1
                        if (project2 != 0) and (project2 != None) and (project2 != '#N/A'):
                            utils.emp_project_update(emp_id,project2,2)

                        else:
                            emp_proj_update = EmployeeProject.objects.filter(emp_id = emp_id,priority = 2,status = 1).update(status = 0)

                        #project3 updation as same of project1
                        if (project3 != 0) and (project3 != None) and (project3 != '#N/A'):
                            utils.emp_project_update(emp_id,project3,3)

                        else:
                            emp_proj_update = EmployeeProject.objects.filter(emp_id = emp_id,priority = 3,status = 1).update(status = 0)

                log.info("--- %s seconds ---projectupdate" % (time.time() - start_time))
                '''
                    Inserting or updating default projects
                '''
                #taking the default projects from constants
                default_projects = []
                for project in DefaultProjects:
                    default_projects.append(utils.strip_value(project.value))
                #checking and updating the projects in projects table
                for project in default_projects:
                    proj_det = Project.objects.filter(name = project)
                    if len(proj_det) == 0:
                        proj = Project(name = project,code = '',status = 1)
                        proj.save()
                    else:
                        Project.objects.filter(name = project).update(status = 1)
                #taking all emails from the excel and updating employee project table if not present
                output = list(map(elem.get, need_columns))
                row_count = worksheet.max_row
                for i in range(row_count + 1):
                    if i > 1:
                        column_details = [utils.strip_value(cell.value) if not isinstance(cell.value,datetime) else cell.value  for cell in worksheet[i]]
                        # print('email',column_details)
                        email = utils.strip_value(column_details[output[0]])
                        emp_det  = Employee.objects.get(email= email)
                        emp_id = emp_det.emp_id
                        for project in default_projects:
                            proj_det = Project.objects.get(name = utils.strip_value(project))
                            project_id = proj_det.id
                            emp_project_det = EmployeeProject.objects.filter(emp_id = emp_id,project_id=project_id)
                            if len(emp_project_det) == 0:
                                emp_proj = EmployeeProject(emp_id = emp_id, project_id=project_id,priority=0, status = 1)
                                emp_proj.save()
                            else:
                                EmployeeProject.objects.filter(emp_id = emp_id,project_id=project_id,status = 0).update(status = 1)
                log.info("--- %s seconds ---default project update" % (time.time() - start_time))
                #inserting excel file details and saving emp details in mis_info
                mis_info = {'Added Employees': added_emp_info,'Deleted Employees': deleted_emp_info,'Old Employees Activated':old_emails}
                mis = MisInfo(mis_filename = filename,info = mis_info,status =1)
                mis.save()
                '''
                if len(added_emp_info) > 0:
                    template = get_template('welcome.html')
                    for eachemp in added_emp_info:
                        emp_list = list(Employee.objects.filter(email = eachemp).values())
                        exp_time = int((datetime.now() + timedelta(hours=settings.FORGOT_PASSWORD_EXP_TIME)).timestamp())
                        userDetails =  {}
                        userDetails['email'] = eachemp
                        userDetails['type'] = 'forgotpassword'
                        userDetails['datetime'] = exp_time
                        conf_token = utils.encrypt(json.dumps(userDetails))
                        log.info(eachemp+" Welcome Email token: "+conf_token)
                        ctx={
                            "token_url":settings.UI_URL+"reset-password/?token="+conf_token,
                            "name":emp_list[-1]['emp_name'],
                        }
                        mail_content = template.render(ctx)
                        # check_status = send_mail(MailConfigurations.Welcome.value, mail_content, settings.EMAIL_FROM, [eachemp], html_message=mail_content)
                        check_status = 1
                        if (check_status == 1):
                            status = 1
                        else:
                            status = 0
                        emp_id = emp_list[-1]['emp_id']
                        try:
                            added = WelcomeEmailNotification(emp_id = emp_id,status = status)
                            added.save()
                        except Exception as e:
                            log.info(eachemp+" Issue while saving in WelcomeEmailNotification")
                '''
                #------------------------------
                if len(added_emp_info) > 0:
                    template = get_template('welcome.html')
                    for eachemp in added_emp_info:
                        # emp_list = list(Employee.objects.filter(email = eachemp).values())
                        # exp_time = int((datetime.now() + timedelta(hours=settings.FORGOT_PASSWORD_EXP_TIME)).timestamp())
                        # userDetails =  {}
                        # userDetails['email'] = eachemp
                        # userDetails['type'] = 'forgotpassword'
                        # userDetails['datetime'] = exp_time
                        # conf_token = utils.encrypt(json.dumps(userDetails))
                        # log.info(eachemp+" Welcome Email token: "+conf_token)
                        # ctx={
                        #     "token_url":settings.UI_URL+"reset-password/?token="+conf_token,
                        #     "name":emp_list[-1]['emp_name'],
                        #     "email":emp_list[-1]['email'],
                        #     "UI_URL":settings.UI_URL,
                        # }
                        # mail_content = template.render(ctx)
                        # check_status = send_mail(MailConfigurations.Welcome.value, mail_content, settings.EMAIL_FROM, [eachemp], html_message=mail_content)
                        #------------------- new code for welcom email ----------
                        emp_id__ = Employee.objects.get(status=1,email=eachemp)
                        email_service.sendWelcomeMail(emp_id__.emp_id)

                        # check_status = 0
                        # global_email_access = GlobalAccessFlag.objects.filter(status=1,access_type__iexact='EMAIL')
                        # if(len(global_email_access)>0):
                        #     accessed_managers = list(map(lambda x:utils.strip_value(x.emp_name),Employee.objects.filter(role_id=4,status=1)))
                        # else:
                        #     accessed_managers = list(map(lambda x:utils.strip_value(x.emp.emp_name),EmailAccessGroup.objects.filter(status=1)))

                        # if (fun_owner !=0) and (fun_owner != None) and (fun_owner != '#N/A'):
                        #     if(utils.strip_value(fun_owner) in accessed_managers):
                        #         # ret_val = 0
                        #         try:
                        #             if(settings.SENDEMAILTOALL):
                        #                 check_status = send_mail(MailConfigurations.Welcome.value, mail_content, settings.EMAIL_FROM, [eachemp], html_message=mail_content)
                        #                 log.info("MAIL SENT TO {}".format([eachemp]))
                        #                 log.info("Welcome EMAIL NOTIFICATION DATA SENT TO {} SUCCESSFULLY".format(eachemp))
                                        
                        #             else:
                        #                 if(eachemp in settings.CUSTOM_EMAILS):
                        #                     check_status = send_mail(MailConfigurations.Welcome.value, mail_content, settings.EMAIL_FROM, [eachemp], html_message=mail_content)
                        #                     log.info("MAIL SENT TO {}".format([eachemp]))
                        #         except Exception as e:
                        #             log.error(traceback.format_exc())
                        #     else:
                        #         log.error("WELCOME EMAIL NOT SENT TO {} BECAUSE CORRESPONDING MANAGER SHOULD NOT HAVE EMAIL ACCESS TO SEND".format(eachemp))

                        #----------------end of new code for welcome email----------        
                        # check_status = 1
                        # if (check_status == 1):
                        #     status = 1
                        # else:
                        #     status = 0
                        # emp_id = emp_list[-1]['emp_id']
                        # try:
                        #     added = WelcomeEmailNotification(emp_id = emp_id,status = check_status)
                        #     added.save()
                        # except Exception as e:
                        #     log.error(traceback.format_exc())
                        #     log.info(eachemp+" Issue while saving in WelcomeEmailNotification")
                #------------------------------
                print("--- %s seconds ---mis table update" % (time.time() - start_time))
                return Response(utils.StyleRes(True,'Excel successfully processed and updated the details',mis_info),status = 200)

            else:
                #sending back Email ids are dublicate
                error_msg = []
                if (is_emails_unique == False):
                    error_msg.append({'Emails Unique' : 'NO, Check the emails'})

                if (is_empname_unique == False):
                    error_msg.append({'Empname Unique' : 'NO, Check the emp names'})

                if (is_empid_unique == False):
                    error_msg.append({'Staffno Unique' : 'No, Check the Staffnos'})

                if (len(empty_empnames) > 0):
                    error_msg.append({'Empty Empnames' : len(empty_empnames) + ' Emp names are empty or zero'})

                if (len(empty_staffnos) > 0):
                    error_msg.append({'Empty Staffnos' : len(empty_staffnos) + ' Staffnos are empty or zero'})

                if (len(missing_man_list) >0):
                    missing = ','.join(str(x) for x in missing_man_list)
                    error_msg.append({'Manager List' : missing + ' Manger Name missing in the Employees'})

                if (len(invalid_emails) > 0):
                    emails = ','.join(str(x) for x in invalid_emails)
                    error_msg.append({'Invalid Emails' : emails})

                if len(empty_date_of_joinings) > 0:
                    error_msg.append({'Invalid date of joining' : 'Date of joining fields should be instance of date and non empty','Invalid Dates':''.join(map(str,empty_date_of_joinings))})

                if len(empty_locations) > 0:
                    error_msg.append({'Invalid Location' : 'Location should be non empty and should be existed ','Locations': ','.join(map(str,empty_locations))})

                if len(empty_genders) > 0:
                    error_msg.append({'Invalid gender' : 'Gender should be non empty and should be Male, Female, Other','Genders':','.join(map(str,empty_genders))})

                if len(empty_marital_statuses) > 0:
                    error_msg.append({'Invalid Marital Status' : 'Marital Status should be non empty and should be Married, Unmarried','Marital Status':','.join(map(str,empty_marital_statuses))})


                if len(empty_categories) > 0:
                    error_msg.append({'Invalid Designation' : 'Designation should be non empty and should be existed','Designation':','.join(map(str,empty_categories))})
                if(len(same_projects) > 0):
                    error_msg.append({'Employee project names shoud be unique.' : same_projects})
                return Response(utils.StyleRes(False,'Wrong Excel Details'  ,error_msg), status=StatusCode.HTTP_UNPROCESSABLE_ENTITY)
                # if (is_emails_unique == False) or (is_empname_unique == False) or (is_empid_unique == False) or len(empty_empnames) > 0 or len(empty_staffnos) > 0:
                #     return Response(utils.StyleRes(False,'Email ids and Employee Names and Staff nos. should be unique and not empty, 0 and #N/A',{}), status=StatusCode.HTTP_UNPROCESSABLE_ENTITY)
                # elif len(missing_man_list) > 0:
                #     missing = ','.join(str(x) for x in missing_man_list)
                #     return Response(utils.StyleRes(False,'Manger Name missing in the Employees',{missing}), status=StatusCode.HTTP_UNPROCESSABLE_ENTITY)
                # else:
                #     #sending back Email ids are not valid
                #     emails = ','.join(str(x) for x in invalid_emails)
                #     return Response(utils.StyleRes(False,'Bad Email Details:'  ,{emails}), status=StatusCode.HTTP_UNPROCESSABLE_ENTITY)
        else:
            #sending back if excel doesnot have required columns
            error_msg = []
            if is_column_headings_unique == False:
                error_msg.append({'Column headings shoud be unique, please check the column headings'})
            if len(main_list) > 0:
                columns = ','.join(str(x) for x in main_list)
                error_msg.append({'missing_columns':main_list})
            return Response(utils.StyleRes(False,'Wrong Excel Details',error_msg), status=StatusCode.HTTP_UNPROCESSABLE_ENTITY)

class DownloadMIS(APIView):
    @jwttokenvalidator
    @custom_exceptions
    def get(self,request):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        email=auth_details['email']
        if(email not in settings.ADMINS_TO_ACCESS_REPORTS):
            return Response(utils.StyleRes(False,"Unauthorized User"), status=StatusCode.HTTP_UNAUTHORIZED)
        disable = request.GET.get('disable', "false")
        e_date = date.today()
        s_date = date(e_date.year, e_date.month, 1)
        startdate = request.GET.get('startdate', s_date)
        enddate = request.GET.get('enddate', e_date)
        if (startdate == "null" or startdate == None) or (enddate == "null" or enddate == None)   :
            startdate = s_date
            enddate = e_date
        data = {'startdate' : startdate, 'enddate' : enddate}
        serialize_data_date = MisDonloadDisableWithDate(data=data)
        if (not serialize_data_date.is_valid()):
            return Response(utils.StyleRes(False,str(serialize_data_date.errors)), status=StatusCode.HTTP_BAD_REQUEST)
        if(disable=="false" or disable.lower() != "true"):
            isdisable=False
        else:
            isdisable=True 
        response=utils.contentTypesResponce('xl',"MIS"+"_"+str(datetime.strftime(datetime.now().date(), '%d%m%Y'))+".xlsx")
        e=ExcelServices(response,in_memory=True,workSheetName="MIS")
        columns=["Status","Company", "Vendor", "Work Location", "Staff No.", "Name", "Qual","Designation", "DOJ","Marital Status", "Gender", "Actual Project 1", "Code", "Actual Project 2", "Code", "Actual Project 3", "Code", "Billing", "Customer 1", "Customer 2", "Business Segment", "Group", "Domain", "Functional Owner", "Manager's Manager", "Reporting Manager", "Email", "Relieved Date"]
        # emp_data = Employee.objects.prefetch_related('profile').allenabledisableemployee().filter(Q(status = 1) | Q(status = 0 + int(not isdisable), relieved__range = [startdate, enddate])).annotate(
        #         category=F('profile__category__name'),marital_status = Case(When(profile__is_married=1,then=V('Married')),default=V('Unmarried'),output_field=CharField()),
        #         gender = Case(When(profile__gender_id=1,then=V('Male')),When(profile__gender_id=2,then=V('Female')),default=V('Other'),output_field=CharField()),
        #         location=F('profile__location__name'),
        #         doj=F('profile__date_of_join'),
                    
        #         ).order_by('-status')
        if(isdisable):
            emp_data = Employee.objects.prefetch_related('profile').allenabledisableemployee().filter(Q(status = 1) | Q(status = 0, relieved__range = [startdate, enddate])).annotate(
                category=F('profile__category__name'),marital_status = Case(When(profile__is_married=1,then=V('Married')),default=V('Unmarried'),output_field=CharField()),
                gender = Case(When(profile__gender_id=1,then=V('Male')),When(profile__gender_id=2,then=V('Female')),default=V('Other'),output_field=CharField()),
                location=F('profile__location__name'),
                doj=F('profile__date_of_join'),
                    
                ).order_by('-status')
        else:
            emp_data = Employee.objects.prefetch_related('profile').allenabledisableemployee().filter(Q(status = 1)).annotate(
                    category=F('profile__category__name'),marital_status = Case(When(profile__is_married=1,then=V('Married')),default=V('Unmarried'),output_field=CharField()),
                    gender = Case(When(profile__gender_id=1,then=V('Male')),When(profile__gender_id=2,then=V('Female')),default=V('Other'),output_field=CharField()),
                    location=F('profile__location__name'),
                    doj=F('profile__date_of_join'),   
                )
        resp=[columns]
        project_style_format_list = {}
        for each in emp_data:
            managers = {str(empl.priority):empl.manager.emp_name for empl in each.emp.all()}
            projects = {str(proj.priority):proj.project.name for proj in each.employeeproject_set.filter(~Q(project__name__in = DefaultProjects.list()),Q(status=1))}

            ''' If the project count is less than 3 then checking the inactive projcet for date range for each employee  and pull the data'''           

            project_style_format= [False, False, False]
            if(len(projects) < 3):
                emp_proj_ids = { emp_proj.id for emp_proj in each.employeeproject_set.filter(~Q(project__name__in = DefaultProjects.list()),Q(status=0))}
                if(len(emp_proj_ids) > 0):
                    time_tracker_data_list = []
                    ''' filter the data based on emp_proj_id from employee_time_tracker '''

                    for emp_proj_id in emp_proj_ids:
                        
                        time_tracker_data = EmployeeProjectTimeTracker.objects.filter(employee_project_id = emp_proj_id, work_date__gte = startdate, work_date__lte = enddate).aggregate(total_minutes=Coalesce(Sum('work_minutes'),0))
                    
                        if(time_tracker_data['total_minutes'] > 0):
                            project_data = EmployeeProject.objects.filter(id = emp_proj_id).values('project').annotate(project_name = F('project__name'))
                            
                            project_name = project_data[0]['project_name']
                            time_tracker_data['project_name'] = project_name
                            time_tracker_data_list.append(time_tracker_data)
                            
                        else:
                            continue
                    if(len(time_tracker_data_list) >0):
                        sorted_list = []
                        for sorted_data in sorted(time_tracker_data_list, key=operator.itemgetter("total_minutes"), reverse=True):
                            sorted_list.append(sorted_data)
                        sorted_list = sorted_list[:3]
                        project_len = len(projects)
                        for i in range(len(sorted_list)):   
                            if(project_len > 0 and project_len <2):
                                    projects[str(2)] = sorted_list[i]['project_name']
                                    project_style_format[1] = True
                            elif(project_len > 1 and project_len <3):
                                projects[str(3)] = sorted_list[i]['project_name']
                                project_style_format[2] = True
                            else:
                                projects[str(i +1)] = sorted_list[i]['project_name']
                                project_style_format[i] = True

            project_style_format_list[each.staff_no] = project_style_format
            resp.append([each.category,each.company,"",each.location,each.staff_no,each.emp_name,"","",each.doj,each.marital_status,each.gender,projects['1'] if '1' in projects.keys() else "", "",projects['2'] if '2' in projects.keys() else "","",projects['3'] if '3' in projects.keys() else "","","","","","",each.company,"",managers['3'] if '3' in managers.keys() else "",managers['2'] if '2' in managers.keys() else "", managers['1'] if '1' in managers.keys() else "", each.email, each.relieved.strftime('%Y-%m-%d') if each.status == 0 and isdisable == True and  each.relieved != None else "", each.status])

        e.writeExcel(resp,row_start=0,datetimeColList=[8],customFormat={'num_format':'yyyy-mm-dd','align':'center'},is_inactive_project_exit = project_style_format_list)
        del e
        return response