from vedikaweb.vedikaapi.services import leave_service
from django.db.models.fields import BooleanField
from rest_framework.views import APIView
from rest_framework.response import Response
from django.conf import settings
from openpyxl import load_workbook
from datetime import date
from itertools import groupby

from vedikaweb.vedikaapi.models import EmployeeHierarchy,LeaveType, NewHireMonthTimePeriods, NewHireLeaveConfig, LeaveConfig, Category ,  LeaveBalance, Leave , LeaveRequest,LeaveAccessGroup,Employee,EmployeeProject,EmployeeProjectTimeTracker, LeaveDiscrepancy,Project, LocationHolidayCalendar,TimesheetDiscrepancy, GlobalAccessFlag, EmployeeProfile, LeaveBalanceUploaded, leaveRequestDisable
# Serialisers
from vedikaweb.vedikaapi.serializers import  LeaveDiscrepancySerializer,NewHireLeaveConfigSerializer, LeaveConfigSerializer,  LeaveBalanceSerializer, LeaveRequestSerializer, LeaveDetailsSerializer, LeaveTypeSerializer, UpdateLeaveConfigSerializer,  IdOnlySerializer, EmployeeProjectTimeTrackerSerializer, LeaveBalanceUploadedSerializer

from vedikaweb.vedikaapi.constants import  LeaveDayStatus, LeaveDiscrepancyStatus, LeaveRequestStatus, MaxLeaveDaysForLeaveType, MaxRequestsForLeaveType, StatusCode, DefaultProjects,LeaveMailTypes, LeaveExcelHeadings, TimesheetDiscrpancyStatus
from vedikaweb.vedikaapi.utils import utils
from vedikaweb.vedikaapi.decorators import custom_exceptions, is_admin,jwttokenvalidator, is_manager, query_debugger
from django.db.models import Q,F,Prefetch,CharField, Case, When, Value as V
from django.core.exceptions import MultipleObjectsReturned, ValidationError, ObjectDoesNotExist
from django.db.models import Sum
from vedikaweb.vedikaapi.services.xlsxservice import ExcelServices
#logging
import traceback, json
from datetime import datetime, timedelta
import logging
from rest_framework import generics
from django.db.models import Value, When, Case ,FloatField, Count, Subquery, IntegerField
from django.db.models.functions import Coalesce 
from django.db import IntegrityError, transaction
import re
from vedikaweb.vedikaapi.services.email_service import email_service
from rest_framework import serializers

log = logging.getLogger(__name__)




class LeaveBalanceView(APIView):
    @jwttokenvalidator
    @custom_exceptions
    def get(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        emp_id=auth_details['emp_id']
        is_emp_admin = auth_details["is_emp_admin"]
        emp_admin_priority = auth_details["emp_admin_priority"]


        is_hr = self.request.query_params.get('is_hr','false')
        if(is_hr.lower()=='true' and (is_emp_admin == False or emp_admin_priority != 2)  ):
            return Response(utils.StyleRes(False,"All users' leave balance fetch error" , 
        "User is not hr with priority"
        ),status=StatusCode.HTTP_UNAUTHORIZED)

        balance_year = self.request.query_params.get('year',datetime.today().year)        
        all_emp_leaves_balance = leave_service.get_leave_balance(balance_year, emp_id, is_hr)

        return Response(utils.StyleRes(True,"Leave Balance", 
        # serial_emp_leaves_balance.data
        all_emp_leaves_balance
        ),status=StatusCode.HTTP_OK)

    def post(self,request,*args, **kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        emp_id = auth_details["emp_id"]
        is_emp_admin = auth_details["is_emp_admin"]
        emp_admin_priority = auth_details["emp_admin_priority"]
        if(is_emp_admin == False or emp_admin_priority!=2 ):
            return Response(utils.StyleRes(False,"Mis Update Balance update error", "user is not admin with priority 2"
                    ),status=StatusCode.HTTP_UNAUTHORIZED)

        req_data=json.loads(json.dumps(request.data))
        if(request.data['modifiedValue']%.5==0):
            old_leave_bal = leave_service.get_leave_balance(int(datetime.now().year),req_data["emp"],'false')[0]['outstanding_leave_bal']
            
            req_data.update({'year':datetime.now().year,'month':datetime.now().month,'status':1,'acted_by': 'hr','hr_emp_id':emp_id, 'current_leave_balance':old_leave_bal,'modified_leave_balance':req_data["modifiedValue"]})
            
            # req_data.update({'year':datetime.now().year,'month':datetime.now().month,'status':1,'acted_by': 'hr','hr_emp_id':emp_id})

            serial_leave_bal_data = LeaveBalanceSerializer(data=req_data)
            if(serial_leave_bal_data.is_valid()):
                if(req_data['leave_credits']!= 0):
                    # old_leave_bal = leave_service.get_leave_balance(int(datetime.now().year),serial_leave_bal_data.validated_data["emp"].emp_id,'false')[0]['outstanding_leave_bal']
                    serial_leave_bal_data.save()
                    new_leave_bal = leave_service.get_leave_balance(int(datetime.now().year),serial_leave_bal_data.validated_data["emp"].emp_id,'false')[0]['outstanding_leave_bal']
                    try:
                        email_service.informLeaveBalanceChange(serial_leave_bal_data.validated_data["emp"].emp_id,old_leave_bal,new_leave_bal)
                        log.info("email sent successfully for leave balance change to  emp_id {}".format(serial_leave_bal_data.data["emp"]))
                    except Exception as e:
                        log.error(traceback.format_exc())
                        log.error("email not sent for leave balance change to  emp_id {}".format(serial_leave_bal_data.data["emp"]))
                return Response(utils.StyleRes(True,"Update Balance", "balance updated successfully"
            ),status=StatusCode.HTTP_CREATED)
            else:
                return Response(utils.StyleRes(False,"Update Balance", serial_leave_bal_data.errors
            ),status=StatusCode.HTTP_BAD_REQUEST)
        else:
            return Response(utils.StyleRes(False,"Update Correct Balance", "UPDATED BALANCE SHOULD BE DIVIDABLE BY .5"
            ),status=StatusCode.HTTP_EXPECTATION_FAILED)


class LeaveRequestView(APIView):
    # serializer_class = LeaveRequestSerializer
    @jwttokenvalidator
    @custom_exceptions
    def get(self,request,*args,**kargs):
        auth_details = utils.validateJWTToken(request) 
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        emp_id=auth_details['emp_id']
        qp = request.query_params
        is_hr = json.loads(qp.get('is_hr','false'))
        # is_hr = auth_details['is_emp_admin']
        leave_requests,errors = leave_service.get_leave_requests(qp,emp_id,is_hr)
        if errors:
            # print('errors in get leave requests')
            # # print(errors)
            return Response(utils.StyleRes(False,"Invalid parameters",errors),status=StatusCode.HTTP_OK)
        else:
            if len(leave_requests) == 0:
                return Response(utils.StyleRes(True,'No Leave Application for the filter criteria'),status=StatusCode.HTTP_NO_CONTENT)
            serial_leave_requests = LeaveRequestSerializer(leave_requests,many=True,context={'request': request})
            return Response(utils.StyleRes(True,"Leave Applications",serial_leave_requests.data),status=StatusCode.HTTP_OK)

    @jwttokenvalidator
    @custom_exceptions
    @transaction.atomic
    def post(self,request,*args,**kargs): 
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        emp_id=auth_details['emp_id']
        gender=auth_details['gender']

        req_data = request.data.copy()
        # checking the start date and end date in-between 27th December to 31th December
        currenDate = date.today()
        restrict_apply_leave  = leaveRequestDisable.objects.filter(startdate__lte = currenDate, enddate__gte= currenDate) 
        
        # User can't able to apply leave beacuse date is in-between 27th December to 31th December
        if(len(restrict_apply_leave) >0):
            return Response(utils.StyleRes(False,"You cannot apply leave now.  Contact HR"), status=StatusCode.HTTP_NOT_ACCEPTABLE) 
        # test request script 
        # emp_id=req_data.get('emp_id')
        # gender= EmployeeProfile.objects.get(emp_id=emp_id).gender_id

        # if request leave type is marriage then first save the files
        req_data['uploads_invitation']=""

        req_data.update({"emp":emp_id,"status":0,"requested_by": "emp","manager_comments":"" })

        # STEP: if the leave type is a marriage leave then upload the invitations to the storage and save the paths
        file_names_for_db=[]
        if int(request.data['leave_type']) == 3:
            files = req_data.pop('invitation_files') if req_data.get('invitation_files') else []
            for file in files:
                current_file=file
                # file_name_temp =file.temporary_file_path()
                # print(file_name_temp)
                file_name = current_file.name                
                # validation = FileExtensionValidator(allowed_extensions=['.jpeg','.jpg']
                # TODO:: validate MIME type
                # get the extension which is split from the file name and append time stamp at the end and the extension after that
                file_name_timestamp = '_'+str(int(datetime.now().timestamp()))
                file_extension = re.split('(\..*)$',file_name) 
                file_name_with_timestamp = file_extension[0]+file_name_timestamp+file_extension[1]
                destination_file_name = settings.UPLOAD_INVITATIONS_PATH+file_name_with_timestamp
                # print(destination_file_name)
                # move the file to the storage destination
                try:
                    with open(destination_file_name,'wb+') as destination_file:
                        for chunk in current_file.chunks():
                            destination_file.write(chunk) 
                    file_names_for_db.append(file_name_with_timestamp)
                except BaseException as e:
                    log.error(traceback.format_exc())
                    req_data['uploads_invitation']=""
                    # print(e)
                    return Response(utils.StyleRes(False,"Marriage invitation files could not be stored"), status=StatusCode.SERVER_ERROR) 
                finally:                    
                    req_data.update({'uploads_invitation':json.dumps(file_names_for_db)})
        else:
            req_data['uploads_invitation']=""
        
        serial_leave_request = LeaveRequestSerializer(data=req_data)  
        # print(serial_leave_request)
        # req_data = json.loads(json.dumps(request.data))
        # print(serial_leave_request.is_valid(), serial_leave_request.errors, req_data)
        # serializers.ImageSerializer(data=request.FILES, many=isinstance(request.FILES, list))
        # return Response(utils.StyleRes(True,"Leave Request", "Leave request has been create successfully"), status=StatusCode.HTTP_OK)
        if(serial_leave_request.is_valid()): 
            start_date = datetime.strptime(req_data['startdate'], '%Y-%m-%dT%H:%M:%S')
            end_date = datetime.strptime(req_data['enddate'], '%Y-%m-%dT%H:%M:%S')
            leave_obj = []

            leave_type = int(req_data['leave_type'])
            start_date_to_get_all_dates = datetime.fromtimestamp(start_date.timestamp())
            leave_dates = []

            leave_types = list(LeaveType.objects.all())
            leave_types_dict = dict() 
            for leave_type_item in leave_types:
                leave_types_dict[leave_type_item.id]=leave_type_item.name
                
            leave_name = leave_types_dict.get(leave_type)

            # STEP: validate gender against leave type and send error if not compatible
            if (leave_name == 'Maternity' and gender == 'Male') or (leave_name == 'Paternity' and gender == 'Female'): 
                return Response(utils.StyleRes(False,'Gender conflict with leave type',{}),status=StatusCode.HTTP_BAD_REQUEST)

            # STEP: check based on the leave type id there are leaves available for that leave type ( Marriage, Maternity, Paternity) 
            # prev_leave_request_for_leave_type = LeaveRequest.objects.filter(emp_id=emp_id,leave_type=leave_type).exclude(status = 3)
            # if leave_name == 'Marriage' or leave_name == 'Maternity' or leave_name == 'Paternity':
            #     # check if the employee is has already consumed the allowed no of leave requests
            #     if prev_leave_request_for_leave_type.count() > (MaxRequestsForLeaveType[leave_name].value - 1):
            #         return Response(utils.StyleRes(False,"No "+leave_name+" leaves available"),status=StatusCode.HTTP_NOT_ACCEPTABLE)

            # STEP: find out if there are any dates overlapping with any of the existing leave requests for this employee
            # get all the dates in the selected dates
            
            holiday_obj = LocationHolidayCalendar.objects.getdetailedHolidayList(emp_id=emp_id)
            holiday_list = list(map(lambda x:datetime.strptime(str(x.holiday_date.strftime('%Y-%m-%d %H:%M:%S')),'%Y-%m-%d %H:%M:%S'),holiday_obj))
            while(start_date_to_get_all_dates <= end_date):
                if start_date_to_get_all_dates.weekday()==5 or start_date_to_get_all_dates.weekday()==6 or start_date_to_get_all_dates in holiday_list:
                    if leave_name != 'Maternity':
                        start_date_to_get_all_dates = start_date_to_get_all_dates + timedelta(days=1)
                        continue
                leave_dates.append(start_date_to_get_all_dates)
                start_date_to_get_all_dates = start_date_to_get_all_dates + timedelta(days=1)

  # query builder to check if the date is between start date and end date
            is_date_between_startdate_enddate = (Q(startdate__gte=start_date) & Q(startdate__lte=end_date)) | (Q(enddate__gte=start_date) & Q(enddate__lte=end_date)) | (Q(startdate__lte=start_date) & Q(enddate__gte=end_date))

            # get all the leave_request_ids which are under this employee and are having date greater than or equal to today
            leave_requests_already_there = list(LeaveRequest.objects.filter(Q(emp_id=emp_id) & is_date_between_startdate_enddate).exclude(status__in =[LeaveRequestStatus.Rejected.value,LeaveRequestStatus.EmployeeCancelled.value]).exclude(leavediscrepancy__status=1).values_list('id'))

            LeaveDiscrepancy.objects.filter(leave_request_id__in=leave_requests_already_there)
            # get all the leave days pertaining to the existing leave_requests
            leave_dates_in_existing_not_consumed_obj = Leave.objects.filter(leave_request_id__in=leave_requests_already_there,leave_on__in=leave_dates)

            leave_dates_in_existing_not_consumed_requests = list(leave_dates_in_existing_not_consumed_obj.values_list('leave_on', flat=True))
            
            leave_dates_in_existing_not_consumed_leave_day_types = list(leave_dates_in_existing_not_consumed_obj.values_list('day_leave_type', flat=True))

            # check all the leave days pertaining to the above existing leave_requests for having any clash with the selected dates 
            # print('array intersection')

            leaves_dates_already_existing =  list(set(leave_dates).intersection(set(leave_dates_in_existing_not_consumed_requests)))

            #  if there is a clash then send an error message other wise continue further processing of the http request
            if len(leaves_dates_already_existing) > 0:
                if(req_data['day_leave_type']=='Half Day' and ((leave_dates_in_existing_not_consumed_leave_day_types[0]=='FIRST_HALF' and req_data["hour"]=='SECOND')or(leave_dates_in_existing_not_consumed_leave_day_types[0]=='SECOND_HALF' and req_data["hour"]=='FIRST'))):
                    return Response(utils.StyleRes(False, "It is not recommended to apply for the other half of the day. It is recemmended to cancel the first request and apply for full day leave", leaves_dates_already_existing),status=408)
                return Response(utils.StyleRes(False, "Some of the requested dates are part of another leave request", leaves_dates_already_existing),status=StatusCode.HTTP_CONFLICT)

            # STEP: find out the total number of days requested and based on leave type reject request if the no of days is more than allowed days for  ( Marriage, Maternity, Paternity)
            total_days =len(leave_dates) 
            if req_data['day_leave_type'] == "Multiple Days":
                if req_data.get('start_date_second_half') == 'true':
                    total_days = total_days - 0.5
                if req_data.get('end_date_first_half') == 'true':
                    total_days = total_days - 0.5 
            if leave_name == 'Marriage' or leave_name == 'Maternity' or leave_name == 'Paternity':
                emp_category = EmployeeProfile.objects.get(emp_id=emp_id).category
                max_leaves = LeaveConfig.objects.get(category_id=emp_category,leave_type__name=leave_name,status=1).max_leaves
                if total_days > max_leaves:
                    return Response(utils.StyleRes(False,"Max leave days allowed for "+leave_name+" is "+str(max_leaves)),status=StatusCode.HTTP_NOT_ACCEPTABLE)

             # STEP: find out if the start date is from the past or is equal to today so as to set the status of the leave request to AutoApprovedEmp
            today = datetime.today()
            today = today.replace(hour=0,minute=0,second=0,microsecond=0)
            
            if utils.is_valid_leave_date(start_date,today):
                # print('start date is less than today', LeaveRequestStatus.AutoApprovedEmp.value)
                # serial_leave_request.data['status'] = [LeaveRequestStatus.AutoApprovedEmp.value]
                # if(len(req_data.getlist("time_tracker_id"))>0):
                #     req_data['status']=LeaveRequestStatus.TimesheeDiscrepancy.value
                # else:
                req_data['status'] = LeaveRequestStatus.AutoApprovedEmp.value
                serial_leave_request = LeaveRequestSerializer(data=req_data)
                serial_leave_request.is_valid() 

            # STEP: save the leave request
            serial_leave_request.save()
            # print(LeaveRequestStatus.AutoApprovedEmp)
            log.info("leave requested by employee with emp_id {}".format(serial_leave_request.data['emp']))
            
            # 
            if("Half Day" in req_data["day_leave_type"]):
                if("SECOND" in req_data["hour"]):
                    day_leave_type = "SECOND_HALF"
                else:
                    day_leave_type = "FIRST_HALF"                
            else:
                day_leave_type = "FULL"
                 
            start_date_duplicate = datetime.fromtimestamp(start_date.timestamp()) 
            timesheet_discrepancy = []
            if(len(req_data.getlist("time_tracker_id"))>0):
                for ts_dis,ts_dis_work_minutes in zip(req_data.getlist("time_tracker_id"),req_data.getlist("modified_work_minutes")):
                    timesheet_discrepancy.append(TimesheetDiscrepancy(leave_request_id = serial_leave_request.data["id"],status = TimesheetDiscrpancyStatus.Pending.value,employee_project_time_tracker_id=ts_dis,work_minutes=int(ts_dis_work_minutes)))
                TimesheetDiscrepancy.objects.bulk_create(timesheet_discrepancy)
                log.info("added discrepancy for employee_project_time_tracker_id {}".format(ts_dis))
            # STEP: Save the leave days in the request
            # new
            emp_vacation_project_id=EmployeeProject.objects.get(Q(project__name=DefaultProjects.Vacation.value)&Q(emp=emp_id)).id
            emptt_data = []
            while(start_date <= end_date):             
                if(start_date.weekday()==5 or start_date.weekday()==6 or start_date in holiday_list):
                    if(leave_name=='Maternity'):
                        leave_obj.append(Leave(leave_request_id = serial_leave_request.data["id"],leave_on=start_date,day_leave_type=day_leave_type,status=LeaveDayStatus.Pending.value))
                    pass
                    # print("skipping weekly holiday and holiday")
                else:
                    # for multiple days request check if there is half day at the start or end
                    if req_data['day_leave_type'] == "Multiple Days":
                        if start_date == end_date:
                            day_leave_type = 'FIRST_HALF' if req_data.get('end_date_first_half')   else 'FULL'
                            # print(day_leave_type,'at end date',req_data.get('end_date_first_half'))
                        elif start_date ==  start_date_duplicate:                           
                            day_leave_type = 'SECOND_HALF' if req_data.get('start_date_second_half') else 'FULL'
                            # print(day_leave_type,'at start date',req_data.get('start_date_second_half'))
                        else:
                            day_leave_type = 'FULL'
                    # new
                    emptt_data.append({'employee_project':emp_vacation_project_id,'work_minutes':(480 if day_leave_type=="FULL" else 300),'work_date':start_date.date(),'work_week':start_date.isocalendar()[1],'work_year':start_date.isocalendar()[0],'status':1})
                    leave_obj.append(Leave(leave_request_id = serial_leave_request.data["id"],leave_on=start_date,day_leave_type=day_leave_type,status=LeaveDayStatus.Pending.value))
                start_date = start_date + timedelta(days=1)
            Leave.objects.bulk_create(leave_obj)
            emptt_serial_data = EmployeeProjectTimeTrackerSerializer(data=emptt_data,many=True)
            if(emptt_serial_data.is_valid()):
                emptt_serial_data.save()
                log.info("employee_project_time_tracker vacation data is updated for employee_project id {} ".format(emp_vacation_project_id))


            # new
            # emp_vacation_project_id=EmployeeProject.objects.get(Q(project__name=DefaultProjects.Vacation.value)&Q(emp=emp_id)).id

            # time_tracker_vacation_data = []
            # for each_dis in discrepancy_details:
            #     ep_tt_data = EmployeeProjectTimeTracker.objects.filter(Q(id=each_dis.employee_project_time_tracker_id)&~Q(employee_project__project__name__in=[DefaultProjects.Vacation.value,DefaultProjects.Holiday.value,DefaultProjects.Mis.value]))
            #     ep_tt_data.update(work_minutes=each_dis.work_minutes)
            #     for each_ep_tt_data in ep_tt_data:
            #         time_tracker_vacation_data.append({'employee_project':emp_vacation_project_id,'work_minutes':480,'work_date':each_ep_tt_data.work_date,'work_week':each_ep_tt_data.work_week,'status':1})
            try:
                email_service.sendLeaveMail(serial_leave_request.data["id"],LeaveMailTypes.Applied.value)
                log.info("email sent successfully for the request id".format(serial_leave_request.data["id"]))
            except Exception as e:
                log.error(traceback.format_exc())
                log.error("email not sent for the request id".format(serial_leave_request.data["id"]))

            
            log.info("added leave request in leave table for leave_request_id {}".format(serial_leave_request.data["id"]))
            return Response(utils.StyleRes(True,"Leave Request","Leave request has been create successfully"),status=StatusCode.HTTP_OK)
        return Response(utils.StyleRes(False,"Leave Request",serial_leave_request.errors),status=StatusCode.HTTP_BAD_REQUEST)

    
    @jwttokenvalidator
    @custom_exceptions 
    def patch(self,request,*args,**kargs): 
        pass

# special leave requests (Marriage, Maternity, Paternity) available for a user
class SpecialLeaveRequestsAvailableView(APIView):
    
    @jwttokenvalidator
    @custom_exceptions
    def get(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        emp_id=auth_details['emp_id']
        # STEP: check based on the leave type id there are leaves available for that leave type ( Marriage, Maternity, Paternity)
        leave_types = list(LeaveType.objects.all())
        leave_types_to_check = ['Marriage','Maternity','Paternity']
        leave_requests_available_basedon_leavetype = list()
        emp_category = EmployeeProfile.objects.get(emp_id=emp_id).category
        leave_config = LeaveConfig.objects.prefetch_related('leave_type').filter(category_id=emp_category,status=1)
        # for leave_type_item in leave_types:
        #     if leave_type_item.name in leave_types_to_check:
        #         # prev_leave_request_for_leave_type = LeaveRequest.objects.filter(emp_id=emp_id,leave_type=leave_type_item).exclude(status = LeaveRequestStatus.EmployeeCancelled.value)
            
        #         leave_requests_available_basedon_leavetype.append({
        #             'id':leave_type_item.id,
        #             'name':leave_type_item.name,
        #             'available':MaxRequestsForLeaveType[leave_type_item.name].value - prev_leave_request_for_leave_type.count()
        #         })

        for each_config in leave_config:
            leave_requests_available_basedon_leavetype.append({
                'id':each_config.leave_type_id,
                'name':each_config.leave_type.name,
                'available':each_config.max_leaves
            })

        return Response(utils.StyleRes(True,"Leave request available for leave types",leave_requests_available_basedon_leavetype),status=StatusCode.HTTP_OK)

# To get the  details of a single leave request
class LeaveRequestSingleView(APIView):
    @jwttokenvalidator
    @custom_exceptions
    def get(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request) 
        if(auth_details['email']==""):
            # print(auth_details)
            return Response(auth_details, status=400)

        user_id=auth_details['emp_id']
        role_id=auth_details['role_id']
        # print(role_id)
        is_manager = json.loads(request.query_params.get('is_manager','false')) 
        leaves_in_last_n_days =  json.loads(request.query_params.get('leaves_in_last_n_days','60'))
        get_discrepancy =  json.loads(request.query_params.get('get_discrepancy','false'))
        conditions = dict()
        # necessary to show the previous leaves of this leave type
        conditions_previous=dict()
        id = kwargs['id']
        conditions['id']=id
        if not is_manager:
            conditions_previous['emp_id']=user_id
            conditions['emp_id']=user_id
        
        leave_request_serializer = IdOnlySerializer(data=kwargs)

        if leave_request_serializer.is_valid():
            try:
                leave_requests = LeaveRequest.objects.prefetch_related('leave_set').annotate(
                    day_count = Sum(Case( When(leave__day_leave_type='FULL', then=1.0),
                    When(leave__day_leave_type='FIRST_HALF', then=0.5),
                    When(leave__day_leave_type='SECOND_HALF', then=0.5),
                    default=0.0,output_field=FloatField(),)),
                    req_status = Case( When(status=LeaveRequestStatus.Pending.value, then=Value('Pending')),
                    When(status=LeaveRequestStatus.Approved.value, then=Value('Approved')),
                    When(status=LeaveRequestStatus.AutoApprovedEmp.value, then=Value('AutoApprovedEmp')),
                    When(status=LeaveRequestStatus.AutoApprovedMgr.value, then=Value('AutoApprovedMgr')),
                    When(status=LeaveRequestStatus.Rejected.value, then=Value('Rejected')),
                    When(status=LeaveRequestStatus.EmployeeCancelled.value, then=Value('Cancelled')),output_field=CharField(),),
                    leave_type_name = F("leave_type_id__name"),
                    # leaves= Concat('leave__id',V(',')),
                    emp_name=F('emp_id__emp_name'), 
                ).get(**conditions)  
                discrepancyFound = False
                if get_discrepancy: 
                    try:
                        leave_discrepancy = LeaveDiscrepancy.objects.get(leave_request_id=id) 
                        discrepancyFound = True
                        leave_discrepancy_details= LeaveDiscrepancySerializer(leave_discrepancy)
                    except (MultipleObjectsReturned, ObjectDoesNotExist):
                        # log.error(traceback.format_exc())
                        pass
                # for special leave types we need to add the previous instances of the special leave type if there are any
                emp_id = leave_requests.emp_id
                conditions_previous['leave_type_id__in']=[3,4,5]
                conditions_previous['leave_type_id']=leave_requests.leave_type_id 
                conditions_previous['emp_id']=emp_id
                leave_requests_previous = LeaveRequest.objects.prefetch_related('leave_set').filter(**conditions_previous).exclude(Q(id=id) | Q(status=LeaveRequestStatus.EmployeeCancelled.value)).annotate(
                    day_count = Sum(Case( When(leave__day_leave_type='FULL', then=1.0),
                    When(leave__day_leave_type='FIRST_HALF', then=0.5),
                    When(leave__day_leave_type='SECOND_HALF', then=0.5),
                    default=0.0,output_field=FloatField(),)),
                    req_status = Case( When(status=0, then=Value('Pending')),
                    When(status=LeaveRequestStatus.Approved.value, then=Value('Approved')),
                    When(status=LeaveRequestStatus.Rejected.value, then=Value('Rejected')),
                    When(status=LeaveRequestStatus.AutoApprovedEmp.value, then=Value('AutoApprovedEmp')),
                    When(status=LeaveRequestStatus.AutoApprovedMgr.value, then=Value('AutoApprovedMgr')),
                    When(status=LeaveRequestStatus.EmployeeCancelled.value, then=Value('Cancelled')),output_field=CharField(),),
                    leave_type_name = F("leave_type_id__name"),
                    # leaves= Concat('leave__id',V(',')),
                    emp_name=F('emp_id__emp_name')
                )

                current_year = datetime.today().year
                leave_dates = list(Leave.objects.filter(leave_request_id=id).values())
                emp_id = leave_requests.emp_id
                today = datetime.today()
                today = today.replace(hour=0,minute=0,second=0,microsecond=0)
                today_minus_n = today - timedelta(days=leaves_in_last_n_days)
                leave_valid_statuses = [LeaveRequestStatus.Pending.value,LeaveRequestStatus.Approved.value,LeaveRequestStatus.AutoApprovedEmp.value,LeaveRequestStatus.AutoApprovedMgr.value]

                leave_descripancy_check = Q(leavediscrepancy__status__isnull=True) | Q(leavediscrepancy__status__in=[LeaveDiscrepancyStatus.Rejected.value,LeaveDiscrepancyStatus.Pending.value])
                all_paid_leave_request_this_year_of_emp = list(LeaveRequest.objects.filter(leave_descripancy_check,emp_id=emp_id,enddate__year=current_year,leave_type__name='Paid',status__in=leave_valid_statuses).values_list('id',flat=True))
                
                is_date_between_today_and_minus_n = (Q(startdate__gte=today_minus_n) & Q(startdate__lte=today)) | (Q(enddate__gte=today_minus_n) & Q(enddate__lte=today)) | (Q(startdate__lte=today_minus_n) & Q(enddate__gte=today))

                all_paid_leave_request_last_n_days = list(LeaveRequest.objects.filter(Q(emp_id=emp_id) & is_date_between_today_and_minus_n & Q(status__in=[LeaveRequestStatus.Approved.value,LeaveRequestStatus.AutoApprovedEmp.value,LeaveRequestStatus.AutoApprovedMgr.value])).values_list('id',flat=True))
                
                no_of_leaves_in_last_n_days = Leave.objects.filter(leave_on__lte=today,leave_on__gt=today_minus_n,leave_request_id__in=all_paid_leave_request_last_n_days,leave_request__leave_type__name__in=["Paid"]).aggregate(
                    count=Sum(Case( When(day_leave_type='FULL', then=1.0),
                    When(day_leave_type='FIRST_HALF', then=0.5),
                    When(day_leave_type='SECOND_HALF', then=0.5),
                    default=0.0,output_field=FloatField(),)))['count']
                # print(all_paid_leave_request_last_n_days)
                leave_credits_this_year = LeaveBalance.objects.filter(year=current_year,emp_id=emp_id).aggregate(count=Coalesce(Sum('leave_credits'),V(0)))
                leaves_this_year = Leave.objects.filter(leave_on__year=current_year,leave_request__in=all_paid_leave_request_this_year_of_emp).aggregate(
                    leave_count= Sum(Case(
                        When(day_leave_type='FULL',then=1.0),
                        When(day_leave_type = 'FIRST_HALF',then=0.5),
                        When(day_leave_type="SECOND_HALF",then=0.5),default=0.0,
                    ),output_field=FloatField())
                )

                leave_balance =  leave_credits_this_year['count'] - (leaves_this_year['leave_count']  if leaves_this_year['leave_count']  else 0)

                serial_leave_requests = LeaveRequestSerializer(leave_requests,context={'request':request}) 
                serial_leave_requests_dict = dict(serial_leave_requests.data)
                serial_leave_requests_dict['leaves']=leave_dates 
                serial_leave_requests_dict['leaves_previous']=LeaveRequestSerializer(leave_requests_previous, many=True,context={'request':request}).data
                serial_leave_requests_dict['leave_balance']=leave_balance 
                if get_discrepancy:
                    if discrepancyFound:
                        serial_leave_requests_dict['leave_discrepancy_details']=leave_discrepancy_details.data
                        serial_leave_requests_dict['discrepancy_raised']=True
                    else:
                        serial_leave_requests_dict['discrepancy_raised']=False
                else:
                    serial_leave_requests_dict['discrepancy_raised']=False

                serial_leave_requests_dict['no_of_leaves_in_last_n_days']=no_of_leaves_in_last_n_days

                for leave_req_previous in serial_leave_requests_dict['leaves_previous']:
                    leave_dates_previous = list(Leave.objects.filter(leave_request_id=leave_req_previous['id']).values())
                    leave_req_previous['leaves']=leave_dates_previous
                return Response(utils.StyleRes(True,"Leave Application",serial_leave_requests_dict),status=StatusCode.HTTP_OK)
            except ObjectDoesNotExist as ex:
                log.error(traceback.format_exc())
                return Response(utils.StyleRes(False,'Leave request does not exist'),status=StatusCode.HTTP_NOT_FOUND)
        else:
            return Response(utils.StyleRes(False,"Bad Request",leave_request_serializer.errors),status=StatusCode.HTTP_BAD_REQUEST)
    
    # transaction atomic to rollback if there is database error
    @jwttokenvalidator
    @custom_exceptions
    @transaction.atomic
    def delete(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) # DISCUSS FOR 401
        emp_id=auth_details['emp_id']
        # print(request)
        id = kwargs['id'] 
        leave_request_serializer = IdOnlySerializer(data=kwargs)
        if leave_request_serializer.is_valid(): 
            try:
                leave_request = LeaveRequest.objects.get(emp_id=emp_id, id=id)
                leave_status = leave_request.status

                # if the leave status is rejected then return without cancelling
                if leave_status == LeaveRequestStatus.Rejected.value: 
                    return Response(utils.StyleRes(False,"Rejected leaves cannot be cancelled",""), status=StatusCode.HTTP_PRECONDITION_FAILED)

                today = datetime.today()
                today = today.replace(hour=0, minute=0, second=0, microsecond=0) 

                
                # the employee can cancel the leave anytime before approved or rejected, before leave dates.
                if utils.is_valid_leave_date(leave_request.startdate,today) and utils.is_valid_leave_date(today,leave_request.enddate):
                    leave_dates_relative_to_today = "between"
                elif (today < leave_request.startdate):
                    leave_dates_relative_to_today = 'before'
                elif (today > leave_request.enddate):
                    leave_dates_relative_to_today = 'after'
                
                # employee cannot cancel an approved leave which is in progress
                if leave_status == LeaveRequestStatus.Approved.value and (leave_dates_relative_to_today == 'between' or leave_dates_relative_to_today == 'after'):
                    return Response(utils.StyleRes(False,"Cannot cancel the leave as the leave is approved and the leave is in progress",""), status=StatusCode.HTTP_PRECONDITION_FAILED)
                
                
                if leave_request.status == LeaveRequestStatus.EmployeeCancelled.value:
                    return Response(utils.StyleRes(False,"Leave request already deleted",""), status=StatusCode.HTTP_PRECONDITION_FAILED)
                elif leave_request.enddate < today - timedelta(days=30):
                    return Response(utils.StyleRes(False,"Cannot delete leave request as it is expired",""), status=StatusCode.HTTP_PRECONDITION_FAILED)
                else: 
                    # Not updating Leave request individual status as of now
                    # leave_dates = Leave.objects.filter(leave_request_id=id)
                    leave_request.status = LeaveRequestStatus.EmployeeCancelled.value
                    leave_request.save()
                    emp_vacation_project_id=EmployeeProject.objects.get(Q(project__name=DefaultProjects.Vacation.value)&Q(emp=emp_id)).id
                    request_dates=(list(map(lambda x : x.date(),leave_request.leave_set.all().values_list('leave_on',flat=True))))
                    EmployeeProjectTimeTracker.objects.filter(Q(employee_project_id=emp_vacation_project_id)&Q(work_date__in=request_dates)).update(work_minutes=0)
                    log.info("employee_project_time_tracker data has been updated for employee_project id {} for dates {}".format(emp_vacation_project_id,','.join(map(str,request_dates))))
                    # leave_dates.update(status=LeaveDayStatus.HrCancelled.value)
                    email_service.sendLeaveMail(leave_request.id,LeaveMailTypes.Cancelled.value)
                    return Response(utils.StyleRes(True,"Leave request has been deleted successfully",""),status=StatusCode.HTTP_OK)

            except ObjectDoesNotExist:
                log.error(traceback.format_exc())
                return Response(utils.StyleRes(False,"The leave request is not found",""),status=StatusCode.HTTP_NOT_FOUND) 
            except:
                log.error(traceback.format_exc())
                return Response(utils.StyleRes(False,"Something went wrong",""), status=StatusCode.HTTP_UNPROCESSABLE_ENTITY)
        else:
            return Response(utils.StyleRes(True,"something wrong",leave_request_serializer.errors),status=StatusCode.HTTP_BAD_REQUEST)

# leave config based on employee type
class LeaveConfigOfCategory(APIView):
    @jwttokenvalidator
    @custom_exceptions
    def get(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        category = request.query_params['category']
        class Serializer(serializers.Serializer):
            category=serializers.CharField(required=True)

        serialized_data  = Serializer(data=request.query_params)

        if serialized_data.is_valid():  
            leave_credits = LeaveConfig.objects.filter(Q(category__name=category)).annotate( 
                leave_type_name=F('leave_type__name'), 
                leave_credits=F('max_leaves')).values('id','leave_type_name','leave_credits')
            return Response(utils.StyleRes(True,'success',leave_credits),status=StatusCode.HTTP_OK) 
        else: 
            return Response(utils.StyleRes(False,'invalid input'),status=StatusCode.HTTP_BAD_REQUEST)

# to resolve a leave application to approve / reject.
class LeaveResolveView(APIView):
    @jwttokenvalidator
    @custom_exceptions
    @is_manager
    @transaction.atomic
    def post(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        req_data=request.data
        class LeaveResolveSerializer(serializers.Serializer):
            id = serializers.IntegerField(required=True)
            manager_comments=serializers.CharField( required=False, max_length=1024,allow_blank=True )
            resolution = serializers.IntegerField(required=True)
            is_discrepancy = serializers.CharField(required=True)

            def validate_resolution(self,value):
                allowed_status = [LeaveRequestStatus.Approved.value,LeaveRequestStatus.Rejected.value]
                if value not in allowed_status:
                    raise ValidationError(message="Invalid resolution. Accepted values are 1 and 2")
                return value 

            def validate_is_discrepancy(self,value):
                allowed_values = ['true','false',None]
                if value not in allowed_values:
                    raise ValidationError(message='Invalid values for the is_discrepancy field')
                return value
 

        #  emp_for_request=LeaveRequest.objects.get(id=leave_req_id)
        # emp_vacation_project_id=EmployeeProject.objects.get(Q(project__name=DefaultProjects.Vacation.value)&Q(emp=emp_for_request.emp)).id

        # time_tracker_vacation_data = []
        # for each_dis in discrepancy_details:
        #     ep_tt_data = EmployeeProjectTimeTracker.objects.filter(Q(id=each_dis.employee_project_time_tracker_id)&~Q(employee_project__project__name__in=[DefaultProjects.Vacation.value,DefaultProjects.Holiday.value,DefaultProjects.Mis.value]))
        #     ep_tt_data.update(work_minutes=each_dis.work_minutes)
        #     for each_ep_tt_data in ep_tt_data:
        #         time_tracker_vacation_data.append({'employee_project':emp_vacation_project_id,'work_minutes':480,'work_date':each_ep_tt_data.work_date,'work_week':each_ep_tt_data.work_week,'status':1})


        leave_request_serializer = LeaveResolveSerializer(data=req_data)
        leave_request_text = 'Leave Request' if not req_data['is_discrepancy'] == 'true' else 'Leave discrepancy request'
        if leave_request_serializer.is_valid(): 
            try:
                today = datetime.today()
                today = today.replace(hour=0, minute=0, second=0, microsecond=0)
                leave_request_id = req_data['id']
                if req_data['is_discrepancy'] == 'true':  
                    leave_discrepancy = LeaveDiscrepancy.objects.get(leave_request_id=leave_request_id)
                    leave_discrepancy.status = req_data['resolution']
                    leave_discrepancy.manager_comments = req_data['manager_comments']
                    leave_discrepancy.save()
                    # STEP: change the leave request status also if the resolution is approved                
                    # leave_request = LeaveRequest.objects.get(id = leave_request_id)
                    # leave_request.status = LeaveRequestStatus.DiscrepancyApproved.value if req_data['resolution'] == 1  else LeaveRequestStatus.DiscrepancyRejected.value
                    # leave_request.save()
                    if(req_data['resolution'] == LeaveDiscrepancyStatus.Approved.value):
                        leave_request = LeaveRequest.objects.get(id=leave_request_id)
                        emp_vacation_project_id=EmployeeProject.objects.get(Q(project__name=DefaultProjects.Vacation.value)&Q(emp=leave_request.emp_id)).id
                        request_dates=(list(map(lambda x : x.date(),leave_request.leave_set.all().values_list('leave_on',flat=True))))
                        EmployeeProjectTimeTracker.objects.filter(Q(employee_project_id=emp_vacation_project_id)&Q(work_date__in=request_dates)).update(work_minutes=0)
                        log.info("employee_project_time_tracker data has been updated for employee_project id {} for dates {}".format(emp_vacation_project_id,','.join(map(str,request_dates))))
                        email_service.sendLeaveMail(leave_request_id,LeaveMailTypes.DiscrepancyApproved.value)
                    elif(req_data['resolution'] == LeaveDiscrepancyStatus.Rejected.value):
                        email_service.sendLeaveMail(leave_request_id,LeaveMailTypes.DiscrepancyRejected.value)

                    return Response(utils.StyleRes(True,leave_request_text+" resolved successfully"),status=StatusCode.HTTP_OK)
                else:
                    leave_request = LeaveRequest.objects.get(id=leave_request_id)

                    # STEP: check if the leave is in employee cancelled status and send the response
                    if leave_request.status == LeaveRequestStatus.EmployeeCancelled.value:
                        return Response(utils.StyleRes(False,leave_request_text+" cannot be approved / rejected as the employee has already cancelled the leave"),status=StatusCode.HTTP_PRECONDITION_FAILED)

                    # if the leave resolution is to reject the leave the apply the conditions as per the requirement
                    # only before the leave dates have started the leave can be rejected

                    if req_data['resolution'] == LeaveRequestStatus.Rejected.value and utils.is_valid_leave_date(leave_request.startdate,today):
                        return Response(utils.StyleRes(False,leave_request_text+" cannot be rejected once it is in progress or consumed."),status=StatusCode.HTTP_PRECONDITION_FAILED)
                    
                    leave_request.status = req_data['resolution']
                    leave_request.manager_comments = req_data['manager_comments']
                    leave_request.save()
					# setting vacation hours zero when manager rejects the leave
                    if(req_data['resolution'] == LeaveRequestStatus.Rejected.value):
                        emp_vacation_project_id=EmployeeProject.objects.get(Q(project__name=DefaultProjects.Vacation.value)&Q(emp=leave_request.emp_id)).id
                        request_dates=(list(map(lambda x : x.date(),leave_request.leave_set.all().values_list('leave_on',flat=True))))
                        EmployeeProjectTimeTracker.objects.filter(Q(employee_project_id=emp_vacation_project_id)&Q(work_date__in=request_dates)).update(work_minutes=0)
                        log.info("employee_project_time_tracker data has been updated for employee_project id {} for dates {}".format(emp_vacation_project_id,','.join(map(str,request_dates))))

                    # all_leaves=leave_request.leave_set.all()
                    # emp_vacation_project_id=EmployeeProject.objects.get(Q(project__name=DefaultProjects.Vacation.value)&Q(emp=leave_request.emp)).id
                    # time_tracker_vacation_data = []
                    # for each_leave in all_leaves:
                    #     print("each_leave.day_leave_type ",each_leave.day_leave_type )
                    #     time_tracker_vacation_data.append({'employee_project':emp_vacation_project_id,'work_minutes':480 if each_leave.day_leave_type == 'FULL' else 300,'work_date':each_leave.leave_on.date(),'work_week':each_leave.leave_on.isocalendar()[1],'status':1})
                    # time_tracker_vacation_serial_data=EmployeeProjectTimeTrackerSerializer(data=time_tracker_vacation_data,many=True)
                    # if(time_tracker_vacation_serial_data.is_valid()):
                    #     time_tracker_vacation_serial_data.save()
                    if(req_data['resolution']==LeaveRequestStatus.Rejected.value):
                        email_service.sendLeaveMail(leave_request.id,LeaveMailTypes.Rejected.value)
                    elif(req_data['resolution']==LeaveRequestStatus.Approved.value):
                        email_service.sendLeaveMail(leave_request.id,LeaveMailTypes.Approved.value)

                    return Response(utils.StyleRes(True,leave_request_text+" resolved successfully"),status=StatusCode.HTTP_OK)
            except ObjectDoesNotExist:
                log.error(traceback.format_exc())
                return Response(utils.StyleRes(False,leave_request_text+" not found"),status=StatusCode.HTTP_NOT_FOUND)
        else:
            return Response(utils.StyleRes(False, leave_request_text+"not resolved. Invalid data sent.",leave_request_serializer.errors),status=StatusCode.HTTP_BAD_REQUEST)

# export resolved leaves . filter may be applied
class ExportResolvedLeaves(APIView):
    
    @jwttokenvalidator
    @custom_exceptions
    # @is_manager
    def get(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request) 
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        emp_id=auth_details['emp_id']
        qp = request.query_params
        is_manager = json.loads(qp.get('is_manager','false'))
        is_hr = json.loads(qp.get('is_hr','false'))
        monthly_time_cycle_flag = json.loads(qp.get('cyclewise','false'))
        threshold = self.request.query_params.get('previous',1)
        
        emp_name = ""
        if(qp.get('emp_name') and qp.get('emp_name')!="ALL" ):
            emp_name = str(qp.get('emp_name'))
        else:
            emp_name = Employee.objects.filter(emp_id=emp_id)[0].emp_name
            if(is_hr):
                emp_name = 'All'
            elif(is_manager):
                emp_name = emp_name+"_Team"
        if(monthly_time_cycle_flag):
            res = utils.get_monthly_cycle(date.today(),Threshold=threshold)
            start_date,end_date = str(res[0]),str(res[1])
        else:
            start_date= datetime.strftime(datetime.strptime(qp.get('start_date', datetime.now().strftime('%Y-%m-%dT%H:%M:%S')), '%Y-%m-%dT%H:%M:%S'),'%Y-%m-%d')
            end_date = datetime.strftime(datetime.strptime(qp.get('end_date', datetime.now().strftime('%Y-%m-%dT%H:%M:%S')), '%Y-%m-%dT%H:%M:%S'),'%Y-%m-%d')
        excel_file = utils.contentTypesResponce('xl',emp_name+'_LeaveHistory_'+start_date+"_"+end_date+".xlsx")
        excel = ExcelServices(excel_file,in_memory=True,multisheetFlag=True)
        columns = ['Staff No','Name','Applied on','Start Date','End Date','Total Days','Leave Type','Leave Status']
        excel_data= [columns]
        sheet2_columns = ['Staff No', 'Name', 'Total Leaves']
        sheets2_data=[sheet2_columns]
        hr_columns = ['Staff No', 'Name', 'Leave Credits', 'Modified On','Comments']
        hr_data = [hr_columns]
        leave_requests,errors = leave_service.get_leave_requests(qp,emp_id,is_hr)
        year = date.today().year 
        is_download = True
        emp_id_list = leave_service.get_emp_id_list(qp,emp_id,is_hr)
        leave_balances = leave_service.get_leave_added_by_hr(year, emp_id_list) #leave_service.get_leave_balance(year,emp_id,is_hr, is_download)
        # print("Leave blance:",leave_balances)
        if errors:
            return Response(utils.StyleRes(False,"Invalid parameters",errors),status=StatusCode.HTTP_OK)
        else:
            # if len(leave_requests) == 0:
            #     return Response(utils.StyleRes(True,'No Leave Applications available for the filter criteria'),status=StatusCode.HTTP_NO_CONTENT)
            # else:
                # statuses = ['Pending','Approved','Rejected','Cancelled']
                statuses = [s.name for s in LeaveRequestStatus]
                leave_summary_dict={}         
                for leave_balance in leave_balances:
                    hr_data.append([leave_balance['staff_no'], leave_balance['emp_name'], leave_balance['total_leave_bal'], leave_balance['createddate'].strftime("%Y-%m-%d"), leave_balance['comments'] ])
                for leave_request in leave_requests:
                    lr = leave_request
                    leave_type = lr.leave_type_name
                    if(lr.leave_type_name == 'Paid'):
                        leave_type = 'General'
                    if(lr.emp_staff_no in leave_summary_dict):
                        leave_summary_dict[lr.emp_staff_no]['leave_count'] +=  int(lr.day_count)
                    else:
                        leave_summary_dict[lr.emp_staff_no]={'emp_name':lr.emp_name, 'leave_count':lr.day_count}
                    excel_data.append([
                        lr.emp_staff_no,lr.emp_name,str(datetime.strftime(lr.created,'%Y-%m-%d')),str(datetime.strftime(lr.startdate,'%Y-%m-%d')),str(datetime.strftime(lr.enddate,'%Y-%m-%d')),lr.day_count,leave_type, statuses[lr.status]
                    ])
                for key,value in leave_summary_dict.items():
                    sheets2_data.append([key,value['emp_name'],value['leave_count']])

                # excel.defineMultipleWorksheets(['Leave History','Leave Summary', 'Hr Modifications'],[excel_data,sheets2_data, hr_data],leaveFlag=True)
                excel.defineMultipleWorksheets(['Leave History','Leave Summary'],[excel_data,sheets2_data],leaveFlag=True)
                # excel.writeExcel(excel_data,row_start=0)
                # excel.terminateExcelService()
                del excel
                return excel_file


# get no of leaves in the last n days
class getLeavesLastNDaysView(APIView):
    
    @jwttokenvalidator
    @custom_exceptions
    @is_manager
    def get(self, request, *args,**kwargs):
        auth_details = utils.validateJWTToken(request)
        if auth_details['email']=="" or auth_details['role_id'] == 1:
            return Response(auth_details, status=400)
        
        class LeavesInLastNDaysSerializer(serializers.Serializer):
            emp_id = serializers.IntegerField(required=True)
            no_of_days = serializers.IntegerField(required=True, min_value=1)
            
        
        serializer = LeavesInLastNDaysSerializer(data=request.query_params)
        if serializer.is_valid():
            emp_id = request.query_params.get('emp_id', None)
            no_of_days = int(request.query_params.get('no_of_days'))
            today = datetime.today()
            today = today.replace(hour=0,minute=0,second=0,microsecond=0)
            today_minus_n = today - timedelta(days=no_of_days) 
            is_date_between_today_and_minus_n = (Q(leave__leave_on__gte=today_minus_n) & Q(leave__leave_on__lte=today)) 
            # | (Q(enddate__gte=today_minus_n) & Q(enddate__lte=today)) | (Q(startdate__lte=today_minus_n) & Q(enddate__gte=today))

            all_paid_leave_request_last_n_days = list(LeaveRequest.objects.filter(Q(emp_id=emp_id) & is_date_between_today_and_minus_n & Q(status__in=[LeaveRequestStatus.Approved.value,LeaveRequestStatus.AutoApprovedEmp.value,LeaveRequestStatus.AutoApprovedMgr.value]) & Q(leave_type__name__in=["Paid"])).values_list('id',flat=True))
            
            no_of_leaves_in_last_n_days= Leave.objects.filter(leave_on__lte=today,leave_on__gt=today_minus_n,leave_request_id__in=all_paid_leave_request_last_n_days).aggregate(
            count=Sum(Case( When(day_leave_type='FULL', then=1.0),
            When(day_leave_type='FIRST_HALF', then=0.5),
            When(day_leave_type='SECOND_HALF', then=0.5),
            default=0.0,output_field=FloatField(),)))['count']

            if(no_of_leaves_in_last_n_days==None):
                no_of_leaves_in_last_n_days = 0

            return Response(utils.StyleRes(True,'No of leaves in last '+str(no_of_days)+' days for employee with id '+str(emp_id),no_of_leaves_in_last_n_days),status=StatusCode.HTTP_OK)
        else:
            return Response(utils.StyleRes(False,'Something wrong with the inputs',serializer.errors),status=StatusCode.HTTP_BAD_REQUEST)


# To get the dates and details in a leave request
class LeaveRequestLeaveDatesView(APIView):
    @jwttokenvalidator
    @custom_exceptions
    def get(self,request,*args,**kargs): 
        try:
            auth_details = utils.validateJWTToken(request)
            if(auth_details['email']==""):
                return Response(auth_details, status=400)
            emp_id=auth_details['emp_id']
            req_id = self.request.query_params.get('request_id',None)
            if(req_id is not None):
                leave_history = Leave.objects.filter(leave_request__emp=emp_id,leave_request__id=req_id).annotate(leave_type_name = F('leave_request__leave_type__name')) 
                serial_leave_history = LeaveDetailsSerializer(leave_history,many=True).data
                return Response(utils.StyleRes(True,"Leave History",serial_leave_history),status=StatusCode.HTTP_OK)
            else:
                all_leave_history = []
                leave_requests = Leave.objects.filter(leave_request__emp=emp_id).annotate(leave_type_name = F('leave_request__leave_type__name'),
                ).order_by("-leave_on")
                serial_leave_history = LeaveDetailsSerializer(leave_requests,many=True).data

                return Response(utils.StyleRes(True,"Leave History",serial_leave_history),status=StatusCode.HTTP_OK)
        except Exception as e:
            log.error(traceback.format_exc())
            return Response(utils.StyleRes(False,"Leave History",str(e)),status=StatusCode.HTTP_OK)



class LeaveTypeView(APIView):
    @jwttokenvalidator
    @custom_exceptions
    def get(self, request, *args, **kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        leave_type = list(LeaveType.objects.filter(status=1).values())
        return Response(utils.StyleRes(True,"Leave Types",leave_type),status=StatusCode.HTTP_OK)
    @jwttokenvalidator
    @is_admin
    @custom_exceptions
    def post(self, request, *args, **kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        serial_leave_type = LeaveTypeSerializer(data=request.data)
        if(serial_leave_type.is_valid() == True):
            serial_leave_type.save()
            return Response(utils.StyleRes(True,"leave type added successfully", {}),status=StatusCode.HTTP_CREATED)
        return Response(utils.StyleRes(False,"Leave type",serial_leave_type.errors), status = StatusCode.HTTP_BAD_REQUEST )


class NewHireMonthTimePeriodsView(APIView): 
    @jwttokenvalidator
    @custom_exceptions
    def get(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        month_time_periods = list(NewHireMonthTimePeriods.objects.filter(status=1).order_by('start_date','end_date').values('id','start_date','end_date'))
        return Response(utils.StyleRes(True,"Month time periods for the new hire first month leave credit round off",month_time_periods),status=StatusCode.HTTP_OK)

class LeaveConfigView(APIView): 
    @jwttokenvalidator
    @is_admin
    @custom_exceptions
    def get(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        category_leave_credit = LeaveConfig.objects.filter(leave_type__status=1).select_related('category','leave_type').order_by('category__id','leave_type__id')
        category_leave_credit_serializer = LeaveConfigSerializer(category_leave_credit,many=True )   
        if len(category_leave_credit_serializer.data) > 0:
            return Response(utils.StyleRes(True,"Leave credit based on emp type and leave type", category_leave_credit_serializer.data),status=StatusCode.HTTP_OK)
        else:
            return Response(utils.StyleRes(True,"No content available", category_leave_credit_serializer.data),status=StatusCode.HTTP_NO_CONTENT)
        

    @jwttokenvalidator
    @is_admin
    @custom_exceptions
    def patch(self,request,*args, **kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        # repr(LeaveConfigSerializer())
        serialized_data = UpdateLeaveConfigSerializer(data=request.data, many=True )
        if serialized_data.is_valid():  
            ids = [item['id'] for item in request.data]
            all_leave_type_each_category = LeaveConfig.objects.filter(id__in = ids) 
            ids_found=list()
            for index, leave_type_category in enumerate(all_leave_type_each_category):  
                leave_type_category.max_leaves = request.data[index]['max_leaves']
                ids_found.append(leave_type_category.id) 
             
            if len(ids) == len(all_leave_type_each_category):                
                LeaveConfig.objects.bulk_update(all_leave_type_each_category,['max_leaves'])
                return Response(utils.StyleRes(True, "Leave credits for leave types updated successfully"),status=StatusCode.HTTP_OK)
            else:
                # ids_not_found = utils.array_diff(ids,ids_found)
                ids_not_found = list(set(ids).difference(set(ids_found)))
                # print(ids,ids_found,ids_not_found)
                return Response(utils.StyleRes(False, "Some of the leave credit rows do not exist. No update done.",ids_not_found),status=StatusCode.HTTP_NOT_FOUND)
        else:
            errors = serialized_data.errors
            return Response(serialized_data.errors,status=StatusCode.HTTP_BAD_REQUEST)

class FillLeaveConfigView(APIView):    
    @jwttokenvalidator
    @custom_exceptions
    @transaction.atomic
    def get(self,request,*args,**kwargs): 
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        # get all the rows from the emp type
        categories = Category.objects.all()
        # get all the rows from the leave type table
        leave_types=LeaveType.objects.all()
        # for each employee type(category) find out if there is a corresponding row with all the available leave types.collect the id's of the emptype and leave type as a tuple or other to hold them together
        added_emp_type_leave_type = list()
        # for all the collected emptype and leave type add a row in the mapping table with leave_credit zero and status equal to that of employee type 
        for category in categories:
            for leave_type in leave_types: 
                try:
                    leave_credit = LeaveConfig.objects.get(category_id=category.id,leave_type_id=leave_type.id)
                except ObjectDoesNotExist as ex:
                    log.error(traceback.format_exc())              
                    leave_credit = LeaveConfig(leave_type=leave_type,category=category,max_leaves=0,status= (category.status if leave_type.name != 'Unpaid' else 0))
                    leave_credit.save() 
                    if leave_credit.pk:
                        added_emp_type_leave_type.append(('category__'+str(category.id),'leave_type__'+str(leave_type.id), 'leave_credit__'+str(leave_credit.pk) ))
        return Response(utils.StyleRes(True, "Leave credit rows added for all leave types and employee types",added_emp_type_leave_type),status=StatusCode.HTTP_OK)

class FillNewHireLeaveConfigView(APIView):    
    @jwttokenvalidator
    @custom_exceptions
    @transaction.atomic
    def get(self,request,*args,**kwargs): 
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        # get all the rows from the emp type
        categories = Category.objects.all()
        # get all the rows from the leave type table
        time_periods=NewHireMonthTimePeriods.objects.all()
        # for each employee type(category) find out if there is a corresponding row with all the available leave types.collect the id's of the emptype and leave type as a tuple or other to hold them together
        added_emp_type_time_period = list()
        # for all the collected emptype and leave type add a row in the mapping table with leave_credit zero and status equal to that of employee type 
        for category in categories:
            for time_period in time_periods: 
                try:
                    leave_credit = NewHireLeaveConfig.objects.get(category_id=category.id,time_period_id=time_period.id)
                    
                except ObjectDoesNotExist as ex:
                    log.error(traceback.format_exc())
                    leave_credit = NewHireLeaveConfig(time_period=time_period,category=category,round_off_leave_credit=0,status=category.status)
                    leave_credit.save() 
                    if leave_credit.pk:
                        added_emp_type_time_period.append(('category__'+str(category.id),'time_period__'+str(time_period.id), 'round_off_leave_credit__'+str(leave_credit.pk) ))
                        
        # return 200 if success full else return 422
        return Response(utils.StyleRes(True, "Leave credit rows added for all leave types and new hire time periods",added_emp_type_time_period),status=StatusCode.HTTP_OK)

class NewHireLeaveConfigView(APIView):
    @jwttokenvalidator
    @is_admin
    @custom_exceptions
    def get(self,request,*args,**kwargs): 
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        round_off = list(NewHireLeaveConfig.objects.select_related('category','time_period').order_by('category__id','time_period__id')) 
        round_off_serializer = NewHireLeaveConfigSerializer(round_off, many=True)  
        if len(round_off_serializer.data) > 0:
            return Response(utils.StyleRes(True,"Leave credit round off for new hire based on emp type", round_off_serializer.data), status=StatusCode.HTTP_OK)
        else:
            return Response(utils.StyleRes(True,"No content available", round_off_serializer.data),status=StatusCode.HTTP_NO_CONTENT)
    
    @jwttokenvalidator
    @is_admin
    @custom_exceptions
    def patch(self,request,*args, **kwargs):  
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        serialized_data = NewHireLeaveConfigSerializer(data=request.data,many=True, partial=True) 
        # print(serialized_data.errors, request.data)
        if serialized_data.is_valid(): 
            ids = [item['id'] for item in request.data]
            all_time_periods_each_category = NewHireLeaveConfig.objects.filter(id__in = ids)
            ids_found=list()
            for index, time_period_category in enumerate(all_time_periods_each_category):
                time_period_category.round_off_leave_credit = request.data[index]['round_off_leave_credit']
                ids_found.append(time_period_category.id)
            
            if len(ids) == len(all_time_periods_each_category):                
                NewHireLeaveConfig.objects.bulk_update(all_time_periods_each_category,['round_off_leave_credit'])
                return Response(utils.StyleRes(True,"Leave credits for time periods of this emp type updated successfully"),status=StatusCode.HTTP_OK)
            else:
                ids_not_found = list(set(ids).difference(set(ids_found)))
                return Response(utils.StyleRes(False, "Some of the leave credit rows do not exist. No update done.",ids_not_found),status=StatusCode.HTTP_NOT_FOUND)
            
        else:
            errors = serialized_data.errors
            return Response(serialized_data.errors,status=StatusCode.HTTP_BAD_REQUEST)
        
class LeaveDiscrepancyView(APIView): 
    @jwttokenvalidator
    @custom_exceptions
    def get(self,request,*args,**kwargs):        
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        user_id=auth_details['emp_id']
        # STEP: get all the employees under this user
        employees = EmployeeHierarchy.objects.filter(priority=1,manager_id=user_id).values_list('emp_id',flat=True)

        # STEP: get the list of leave_discrepancies where the leave_request__emp_id is in the above employees list
        # leave_discrepancies = LeaveDiscrepancy.objects.select_related('leave_request','leave_request__emp').prefetch_related(
        #     Prefetch('leave_request',queryset=Leave.objects.filter(leave_request_id=OuterRef(F('leave_request__id'))),to_attr='leaves')
        # ).filter(leave_request__emp_id__in=employees,status=LeaveDiscrepancyStatus.Pending.value).annotate(
        #     emp_name= F('leave_request__emp__emp_name'),
        #     startdate=F('leave_request__startdate'),
        #     enddate=F('leave_request__enddate'),
        #     leave_type_name=F('leave_request__leave_type__name')
        # )
        
        leave_discrepancies = LeaveRequest.objects.prefetch_related(
            Prefetch('leavediscrepancy',queryset=LeaveDiscrepancy.objects.all() ),
            Prefetch('leave_set',queryset=Leave.objects.all() )
        ).filter(emp_id__in=employees,leavediscrepancy__status=LeaveDiscrepancyStatus.Pending.value).annotate( 
            day_count = Sum(Case( When(leave__day_leave_type='FULL', then=1.0),
            When(leave__day_leave_type='FIRST_HALF', then=0.5),
            When(leave__day_leave_type='SECOND_HALF', then=0.5),
            default=0.0,output_field=FloatField(),)),
            emp_name=F('emp__emp_name'),
            emp_staff_no=F('emp__staff_no'),
            leave_type_name=F('leave_type__name'),
            discrepancy_raised=V(True,output_field=BooleanField()),
            status_discrepancy=F('leavediscrepancy__status')
        ).exclude(leavediscrepancy__id=None)
        # print(len(leave_discrepancies1),len(leave_discrepancies), leave_discrepancies1.values(),leave_discrepancies)

        if len(leave_discrepancies) == 0:
            return Response(utils.StyleRes(True,"no leave discrepancies found"),status=StatusCode.HTTP_NO_CONTENT)
        else:            
            serialized_data = LeaveRequestSerializer(leave_discrepancies,many=True,context={'request':request})
            return Response(utils.StyleRes(True,"leave discrepancies list",serialized_data.data),status=StatusCode.HTTP_OK)

    @jwttokenvalidator
    @custom_exceptions 
    def post(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        qp = request.query_params
        leave_request_id = qp.get('leave_request_id')

        # STEP: validate the request for the query params leave_request_id
        if not leave_request_id:
            log.error('leave request id not found')
            return Response(utils.StyleRes(False,'leave_request_id found in query params'),status=StatusCode.HTTP_BAD_REQUEST)

        req_data = request.data
        # STEP: save the leave discrepancy in the db
        try:
            log.debug('leave discrepancy saving')
            leave_discrepancy =   LeaveDiscrepancy(leave_request_id=int(leave_request_id),emp_comments=req_data['emp_comments'])
            leave_discrepancy.save()
            email_service.sendLeaveMail(int(leave_request_id),LeaveMailTypes.DiscrepancyApplied.value)
            return Response(utils.StyleRes(True,'Leave discrepancy created successfully'),status=StatusCode.HTTP_OK)
        except IntegrityError as exception:
            log.error(traceback.format_exc())
            # print('leave discrepancy unprocessable',type(exception).__name__, exception.args)
            if str(exception).find('ui_leave_discrepancy_leave_request') is not -1:
                return Response(utils.StyleRes(True,'Leave discrepancy already exists for this leave request'),status=StatusCode.HTTP_CONFLICT)
            else:
                return Response(utils.StyleRes(False,'Something went wrong'), status = StatusCode.HTTP_UNPROCESSABLE_ENTITY)
        except Exception as ex:
            log.error(traceback.format_exc())
            return Response(utils.StyleRes(False,'Something went wrong'), status = StatusCode.HTTP_UNPROCESSABLE_ENTITY)


class ExportEmpLeave(APIView):
    
    @jwttokenvalidator
    @is_admin
    @custom_exceptions
    def get(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request) 
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        emp_id=auth_details['emp_id']
        qp = request.query_params
        excel_file = utils.contentTypesResponce('xl','All_Employees_LeaveBalance_'+str(datetime.strftime(datetime.now().date(), '%Y-%m-%d'))+".xlsx")
        excel = ExcelServices(excel_file,in_memory=True,workSheetName='LeaveBalance')
        columns = []
        for each in LeaveExcelHeadings:
            columns.append(utils.strip_value(each.value))
        excel_data= [columns]
        leave_requests,errors = leave_service.get_leave_requests(qp,emp_id)
 
        is_hr = self.request.query_params.get('is_hr','false')
        balance_year = self.request.query_params.get('year',datetime.today().year)        
        all_emp_leaves_balance = leave_service.get_leave_balance(balance_year, emp_id, 'true')
        # return Response(utils.StyleRes(True,"Leave Balance", 
        # # serial_emp_leaves_balance.data
        # all_emp_leaves_balance
        # ),status=StatusCode.HTTP_OK)

        if errors:
            return Response(utils.StyleRes(False,"Invalid parameters",errors),status=StatusCode.HTTP_CONFLICT)
        else:
            for emp_leaves_balance in all_emp_leaves_balance:
                # lr = leave_request
                excel_data.append([
                    # lr.emp_name,str(datetime.strftime(lr.startdate,'%d-%m-%Y')),str(datetime.strftime(lr.enddate,'%d-%m-%Y')),lr.day_count,lr.leave_type_name, statuses[lr.status],lr.leave_reason,str(datetime.strftime(lr.updated,'%d-%m-%Y'))
                emp_leaves_balance['emp_name'],emp_leaves_balance['staff_no'],emp_leaves_balance['outstanding_leave_bal']
                ])
            excel.writeExcel(excel_data,row_start=0)
            excel.terminateExcelService()
            del excel
            return excel_file
    # from .decorators import query_debugger 
    # @query_debugger

    @jwttokenvalidator
    @is_admin
    @custom_exceptions
    def post(self,request,*args, **kwargs):
        auth_details = utils.validateJWTToken(request) 
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        emp_id=auth_details['emp_id']
        is_emp_admin = auth_details["is_emp_admin"]
        emp_admin_priority = auth_details["emp_admin_priority"]
        if(is_emp_admin == False or emp_admin_priority!=2 ):

            return Response(utils.StyleRes(False,"Mis Update Balance update error", "user is not admin with priority 2"
                    ),status=StatusCode.HTTP_UNAUTHORIZED)

        excel_name = request.data
        log.debug('excel_name',excel_name)
        #getting excel and updating the excel name with cuttent time stamp and saving excel in directory
        excel_file_name = excel_name['file']
        uploadedfilename = str(excel_file_name).split('.')[0]
        uploadedfileext=str(excel_file_name).split('.')[-1]
        filename=uploadedfilename+"_"+utils.getUniqueId() + '.'+ uploadedfileext
        utils.createDirIfNotExists(settings.UPLOAD_LEAVE_BALANCE_PATH)
        dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(settings.UPLOAD_LEAVE_BALANCE_PATH+filename, 'wb') as f:
            f.write(excel_file_name.read())
        #Getting the sheet names from excel
        wb = load_workbook(excel_name['file'])
        sheet_names = wb.sheetnames
        #taking the first sheet in excel
        worksheet = wb[sheet_names[0]]
        # print('worksheet',len(worksheet[1]))
        # print('max_column_length',worksheet.max_row)
        empty_rows = []
        #getting empty rows deleting from worksheet
        for idx, row in enumerate(worksheet.iter_rows(max_col=worksheet.max_column), start=1):
            empty = not any((utils.strip_value(cell.value) for cell in row))
            if empty:
                empty_rows.append(idx)
        for row_idx in reversed(empty_rows):
            worksheet.delete_rows(row_idx, 1)

        #case insensitive for the headings
        for each in worksheet[1]:
            if each.value != None and type(each.value) != int:
                each.value = each.value.lower()
            else:
                each.value
         #needed columns form excel
        need_columns =[]
        

        for each in LeaveExcelHeadings:
            need_columns.append(utils.strip_value(each.value))

        #taking all excel columns
        excel_columns=[]
        for cell in worksheet[1]:
            if utils.strip_value(cell.value) != None:
                excel_columns.append(cell.value)
            else:
                 excel_columns.append(cell.value)
        log.debug('excel_columns: '+",".join(map(str,excel_columns)))

        row_count = worksheet.max_row #max_row count from excel
        #checking all needed columns are present in excel
        main_list = list(set(need_columns) - set(excel_columns))
        #is headings unique
        is_column_headings_unique  = False
        is_column_headings_unique = (len(set(excel_columns)) == len(excel_columns))

        if(len(main_list)>0 or is_column_headings_unique==False):
            return Response(utils.StyleRes(False,"Mis Update Balance error", {'missing_columns':main_list,'unique_columns':is_column_headings_unique}
        ),status=StatusCode.HTTP_BAD_REQUEST)
        #taking all the emails from excel Email row columns and validating all the emails

        valid_empnames = []
        valid_staffnos = []

        empty_empnames = []
        empty_staffnos = []

        invalid_empnames =[]
        duplicated_empnames = []
        duplicated_staff_no = []
        valid_leave_bals =[]
        valid_modified_leave_bal = []
        mismatched_leave_bals = []
        leave_updated_emplist = []
        leave_bal_data = []
        invalid_leave_data_emplist = []
        empty_comments = []
        all_send_email_list = []
        if len(main_list) == 0:
            log.debug("Yes, list1 contains all elements in list2")
            #checking same project names for employee
            same_projects = []

            elem = {k: i for i, k in enumerate(excel_columns)}
            output = list(map(elem.get, need_columns))
            row_count = worksheet.max_row
            for i in range(row_count + 1):
                if i > 1:
                    column_details = [cell.value for cell in worksheet[i]]
                    # print('email',column_details)
                    emp_name = utils.strip_value(column_details[output[0]])
                    staff_no = utils.strip_value(column_details[output[1]])
                    leave_bal = column_details[output[2]]
                    modified_leave_bal = column_details[output[3]]
                    
                    comments = column_details[output[4]]
                    
                    if(emp_name not in valid_empnames):
                        valid_empnames.append(emp_name)
                    else:
                        duplicated_empnames.append(emp_name)

                    if(staff_no not in valid_staffnos):
                        valid_staffnos.append(staff_no)
                    else:
                        duplicated_staff_no.append(staff_no)
                    valid_leave_bals.append(leave_bal)
                    valid_modified_leave_bal.append(modified_leave_bal)

                    emp_details = Employee.objects.filter(Q(emp_name=emp_name)&Q(staff_no=staff_no)&Q(status=1))
                    if(len(emp_details)==0):
                        invalid_empnames.append(emp_name)
                    else:
                        

                        # print("existing_leave_details",existing_leave_details)
                        if(modified_leave_bal==None):
                            continue
                        if((isinstance(modified_leave_bal,float)==False and isinstance(modified_leave_bal,int)==False) or modified_leave_bal%.5!=0):
                            invalid_leave_data_emplist.append(emp_name)
                            continue
                        if(comments==None or len(comments.strip())==0):
                            empty_comments.append(emp_name)
                            continue

                        existing_leave_details = leave_service.get_leave_balance(datetime.today().year, emp_details[0].emp_id, 'false')[0]
                        if(leave_bal!=existing_leave_details['outstanding_leave_bal']):
                            mismatched_leave_bals.append({'name':emp_name,'leave_bal':existing_leave_details['outstanding_leave_bal']})


                        leave_credits =modified_leave_bal -  existing_leave_details['outstanding_leave_bal']
                        leave_bal_data.append({'emp':emp_details[0].emp_id,'year':datetime.today().year,'month':datetime.today().month,
                        'acted_by':'hr','comments':comments,'status':1,'leave_credits':leave_credits,'hr_emp_id':emp_id
                        })
                        leave_updated_emplist.append(emp_name)
                        all_send_email_list.append([emp_details[0].emp_id,existing_leave_details['outstanding_leave_bal'],modified_leave_bal])
            
            res = {'duplicated_empnames':duplicated_empnames, 'duplicated_empnames':duplicated_empnames,'invalid_empnames':invalid_empnames,
                'mismatched_leave_bals':mismatched_leave_bals, 'invalid_leave_data_emplist':invalid_leave_data_emplist,'empty_comments': empty_comments
            }

            if(len(invalid_empnames)>0 or len(duplicated_empnames)>0 or len(duplicated_staff_no)>0 or 
            len(mismatched_leave_bals)>0 or len(invalid_leave_data_emplist)>0) or len(empty_comments)>0:
                return Response(utils.StyleRes(False,"Mis Update Balance error", res
        ),status=StatusCode.HTTP_BAD_REQUEST)
            
            serial_leave_bal_data = LeaveBalanceSerializer(data=leave_bal_data,many=True)
            leave_balance_uploaded_serial_data = LeaveBalanceUploadedSerializer(data = {'emp': emp_id,'leave_balance_filename':filename,'status':1})

            if(serial_leave_bal_data.is_valid()):
                serial_leave_bal_data.save()
                res.update({'leave_updated_emplist':leave_updated_emplist})
                if(leave_balance_uploaded_serial_data.is_valid()):
                    leave_balance_uploaded_serial_data.save()
                    log.info("uploaded leave balance excel by emp {} filename ".format(emp_id,filename))
                for each_mail in all_send_email_list:
                    try:
                        email_service.informLeaveBalanceChange(each_mail[0],each_mail[1],each_mail[2],True)
                        log.info("email sent successfully for leave balance change to  emp_id {}".format(each_mail[0]))
                    except Exception as e:
                        log.error(traceback.format_exc())
                        log.error("email not sent for leave balance change to  emp_id {}".format(each_mail[0]))
                
                return Response(utils.StyleRes(True,"Mis Update Balance updated successfully", res
                ),status=StatusCode.HTTP_CREATED)
            else:
                return Response(utils.StyleRes(False,"Mis Update Balance error", serial_leave_bal_data.errors
        ),status=StatusCode.HTTP_BAD_REQUEST)

class LeaveStatusAPI(APIView):
    @jwttokenvalidator
    @custom_exceptions
    def get(self,request):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400)
        leave_flag = False
        emp_id=auth_details['emp_id']
        global_leave_access = GlobalAccessFlag.objects.filter(status=1,access_type__iexact='LEAVE')
        if(len(global_leave_access)>0):
            leave_access_grp_list = list(map(lambda x:x.emp_id,Employee.objects.filter(role_id=4,status=1)))
        else:
            leave_access_grp_obj = LeaveAccessGroup.objects.filter(status=1)
            leave_access_grp_list = list(map(lambda x: x.emp_id,leave_access_grp_obj))

        emp_hierarchy_obj = EmployeeHierarchy.objects.filter(manager_id__in=leave_access_grp_list,emp_id=emp_id)
        if(len(emp_hierarchy_obj)>0):
            leave_flag = True
        return Response({'leave_flag':leave_flag})



#BASED ON LEAVE TABLE
#INDIVIDUAL LEAVE DAYS AE CONSIDERING AND RETURNING AS RESPONSE
class MonthyCycleLeaveReportView(APIView):
    @jwttokenvalidator
    @is_admin
    @custom_exceptions
    def get(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        threshold = self.request.query_params.get('previous',1)
        emp_name = self.request.query_params.get('emp_name',None)
        month = self.request.query_params.get('month',None)
        year = self.request.query_params.get('year',None)
        # print(utils.isNotNone(month,year))
        if(utils.isNotNone(month,year)):
            start_and_end_dates = utils.get_start_and_end_dates_based_on_month_year(month,year)
        else:
            start_and_end_dates = utils.get_monthly_cycle(date.today(),Threshold=threshold)
        leaves = Leave.objects.filter(leave_on__gte=start_and_end_dates[0],leave_on__lte=start_and_end_dates[1])
        if(emp_name is not None):
            leaves = leaves.filter(leave_request__emp__emp_name__iexact=emp_name)
        leaves_ = leaves.annotate(
            emp_name = F('leave_request__emp__emp_name')
        ).values('leave_request_id','emp_name','leave_on','day_leave_type','status')
        # define a fuction for key
        def key_func(k):
            return k['leave_request_id']
        
        # sort INFO data by 'company' key.
        LEAVES_INFO = sorted(list(leaves_), key=key_func)
        output = []
        for key, value in groupby(LEAVES_INFO, key_func):
            value=list(value)
            output.append({'leaves_count':len(value),'leave_request_id':key,'values':value})
        return Response(output)

#BASED ON BOTH LEAVE AND LEAVE REQUEST TABLE
#SPLITTING THE LEAVE REQUEST BASED ON MONTHLY CYCLE DATES AND GROUPING CONSECUTIVE LEAVES UNDER THE SAME CYCLE
class MonthyCycleLeaveReportRequestBasedView(APIView):
    @jwttokenvalidator
    @is_admin
    @custom_exceptions
    def get(self,request,*args,**kwargs):
        auth_details = utils.validateJWTToken(request)
        if(auth_details['email']==""):
            return Response(auth_details, status=400) 
        export_flag = self.request.query_params.get('export',False)
        threshold = self.request.query_params.get('previous',1)
        emp_name = self.request.query_params.get('emp_name',None)
        month = self.request.query_params.get('month',None)
        year = self.request.query_params.get('year',None)
        is_future_leave = self.request.query_params.get('is_future_leave',"false")
        start_date= datetime.strptime(self.request.query_params.get('start_date', date.today()), '%Y-%m-%dT%H:%M:%S')
        end_date= datetime.strptime(self.request.query_params.get('end_date', date.today()), '%Y-%m-%dT%H:%M:%S')
        leave_type_filters = [LeaveRequestStatus.Approved.value,LeaveRequestStatus.AutoApprovedEmp.value,LeaveRequestStatus.AutoApprovedMgr.value]

        # print(utils.isNotNone(month,year))

        # if(utils.isNotNone(month,year)):
        #     start_and_end_dates = utils.get_start_and_end_dates_based_on_month_year(month,year)
        # else:
        #     start_and_end_dates = start_date,end_date
        
        
        if(is_future_leave=="true"):
            tomorrow = date.today() + timedelta(days=1)
            year_end_date = date(tomorrow.year, 12, 31)
            start_date= datetime.strptime(str(tomorrow), '%Y-%m-%d')
            end_date= datetime.strptime(str(year_end_date), '%Y-%m-%d')
            # added pending leaves
            leave_type_filters = [LeaveRequestStatus.Approved.value,LeaveRequestStatus.AutoApprovedEmp.value,LeaveRequestStatus.AutoApprovedMgr.value, LeaveRequestStatus.Pending.value]
        else:
            start_and_end_dates = start_date,end_date
        start_and_end_dates = start_date,end_date
        leaves = Leave.objects.filter(Q(leave_on__gte=start_and_end_dates[0]) & Q(leave_on__lte=start_and_end_dates[1]) &Q(leave_request__status__in=leave_type_filters) & Q(leave_request__emp__status=1) & (~Q(leave_request__leavediscrepancy__status = 1)))

        if(emp_name is not None and emp_name.strip()!='' and emp_name.strip().upper()!='ALL'):
            leaves = leaves.filter(leave_request__emp__emp_name__iexact=emp_name)
        else:
            emp_name='ALL'
        leaves_ = leaves.annotate(
            emp_name = F('leave_request__emp__emp_name'),
        ).values('leave_request_id','emp_name','leave_on','day_leave_type','status')

        # fuction for key
        def key_func(k):
            return k['leave_request_id']
        
        # sort INFO data by 'company' key.
        LEAVES_INFO = sorted(list(leaves_), key=key_func)
        output = []
        grouped_leaves = groupby(LEAVES_INFO, key_func)
        keys_list = []
        for key, value in grouped_leaves:
            keys_list.append(key)
        leave_q_details_obj = LeaveRequest.objects.select_related('emp').filter(id__in=keys_list,leave__leave_on__gte=start_and_end_dates[0],leave__leave_on__lte=start_and_end_dates[1],status__in=leave_type_filters,emp__status=1).annotate(
                day_count = Sum(Case( When(leave__day_leave_type='FULL', then=1.0),
                    When(leave__day_leave_type='FIRST_HALF', then=0.5),
                    When(leave__day_leave_type='SECOND_HALF', then=0.5),
                    default=0.0,output_field=FloatField(),)),
                req_status = Case( When(status=0, then=Value('Pending')),
                    When(status=LeaveRequestStatus.Approved.value, then=Value('Approved')),
                    When(status=LeaveRequestStatus.Rejected.value, then=Value('Rejected')),
                    When(status=LeaveRequestStatus.AutoApprovedEmp.value, then=Value('AutoApprovedEmp')),
                    When(status=LeaveRequestStatus.AutoApprovedMgr.value, then=Value('AutoApprovedMgr')),
                    When(status=LeaveRequestStatus.EmployeeCancelled.value, then=Value('Cancelled')),output_field=CharField(),),
                leave_type_name = F("leave_type_id__name"),
                emp_name = F('emp__emp_name'),
                emp_staff_no = F('emp__staff_no'),
                discrepancy_status=F('leavediscrepancy__status')
            )
        leave_req_dict = {}
        for eachleaverequest in leave_q_details_obj:
            leave_req_dict[eachleaverequest.id]=eachleaverequest
        for key, value in groupby(LEAVES_INFO, key_func):
            value=list(value)
            # leave_q_details = LeaveRequest.objects.select_related('emp').filter(id=key,leave__leave_on__gte=start_and_end_dates[0],leave__leave_on__lte=start_and_end_dates[1]).annotate(
            #     day_count = Sum(Case( When(leave__day_leave_type='FULL', then=1.0),
            #         When(leave__day_leave_type='FIRST_HALF', then=0.5),
            #         When(leave__day_leave_type='SECOND_HALF', then=0.5),
            #         default=0.0,output_field=FloatField(),)),
            #     req_status = Case( When(status=0, then=Value('Pending')),
            #         When(status=LeaveRequestStatus.Approved.value, then=Value('Approved')),
            #         When(status=LeaveRequestStatus.Rejected.value, then=Value('Rejected')),
            #         When(status=LeaveRequestStatus.AutoApprovedEmp.value, then=Value('AutoApprovedEmp')),
            #         When(status=LeaveRequestStatus.AutoApprovedMgr.value, then=Value('AutoApprovedMgr')),
            #         When(status=LeaveRequestStatus.EmployeeCancelled.value, then=Value('Cancelled')),output_field=CharField(),),
            #     leave_type_name = F("leave_type_id__name"),
            #     emp_name = F('emp__emp_name'),
            #     emp_staff_no = F('emp__staff_no'),
            # )
            leave_q_details = leave_req_dict[key]
            leave_req_obj = leave_q_details
            day_count = leave_req_obj.day_count
            if('.5' not in str(leave_req_obj.day_count)):
                day_count = int(day_count)
            color_date = None
            if(leave_req_obj.startdate<start_date and leave_req_obj.enddate <= end_date):
                color_date="startdate"
            elif(leave_req_obj.startdate>=start_date and leave_req_obj.enddate > end_date):
                color_date="enddate"
            elif(leave_req_obj.startdate<start_date and leave_req_obj.enddate > end_date):
                color_date="both"
            
            output.append({'emp_name':leave_req_obj.emp_name,'emp':leave_req_obj.emp.emp_id,'colordate':color_date,'day_count':day_count,'id':key,'req_status':leave_req_obj.req_status,'startdate':leave_req_obj.startdate,'enddate':leave_req_obj.enddate,'leave_type_name':leave_req_obj.leave_type_name,'emp_staff_no':leave_req_obj.emp_staff_no,'emp_comments':leave_req_obj.emp_comments,'leave_reason':leave_req_obj.leave_reason,'leave_type':leave_req_obj.leave_type_id,'manager_comments':leave_req_obj.manager_comments,'requested_by':leave_req_obj.requested_by,'applied_on':str(datetime.strftime(leave_req_obj.created,'%d-%m-%Y')),'status':leave_req_obj.status,'uploads_invitation':leave_req_obj.uploads_invitation})
        if(not export_flag):
            return Response(utils.StyleRes(200,"success",output))
        else:
            excel_file = utils.contentTypesResponce('xl',emp_name+'_LeaveHistory_'+str(start_date)+"_"+str(end_date)+"_cycle.xlsx")
            excel = ExcelServices(excel_file,in_memory=True,multisheetFlag=True)
            columns = ['Staff No','Name','Applied on','Start Date','End Date','Total Days','Leave Type','Leave Status']
            excel_data= [columns]
            sheet2_columns = ['Staff No', 'Name', 'Total Leaves']
            sheets2_data=[sheet2_columns]
            leave_summary_dict={}
            
            if len(output) == 0:
                return Response(utils.StyleRes(True,'No Leave Applications available for the filter criteria'),status=StatusCode.HTTP_NO_CONTENT)
            else:
                # statuses = ['Pending','Approved','Rejected','Cancelled']
                statuses = [s.name for s in LeaveRequestStatus]
                for lr in output:
                    if(lr['leave_type_name']=='Paid'):
                        leave_type = 'General'
                    else:
                        leave_type = lr['leave_type_name']
                    excel_data.append([
                        lr['emp_staff_no'],lr['emp_name'],lr['applied_on'],str(datetime.strftime(lr['startdate'],'%d-%m-%Y')),str(datetime.strftime(lr['enddate'],'%d-%m-%Y')),lr['day_count'],leave_type, statuses[lr['status']]
                    ])
                    if(lr['emp_staff_no'] in leave_summary_dict):
                        leave_summary_dict[lr['emp_staff_no']]['leave_count'] +=  lr['day_count']
                    else:
                        leave_summary_dict[lr['emp_staff_no']]={'emp_name':lr['emp_name'], 'leave_count':lr['day_count']}
                for key,value in leave_summary_dict.items():
                    sheets2_data.append([key,value['emp_name'],value['leave_count']])

                excel.defineMultipleWorksheets(['Leave History','Leave Summary'],[excel_data,sheets2_data],leaveFlag=True)
                # excel.writeExcel(excel_data,row_start=0)
                # excel.terminateExcelService()
                del excel
                return excel_file